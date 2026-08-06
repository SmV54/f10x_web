# -*- coding: utf-8 -*-
"""
Consolida Comercial\\_coleta_sindicos.csv na aba da planilha comercial.

Duas regras que existem por causa de erro real encontrado na coleta:

1) UM CONTATO POR PAGINA (a menos que a pagina seja um diretorio conferido
   a mao). Paginas de plataforma (ueniweb, por exemplo) trazem numeros de
   OUTRAS empresas em widgets: a pagina do "Sindico Profissional Joao Pessoa"
   devolveu numeros com DDD 89, 68, 96, 21 e 84 alem do (83) dela. Ficar so
   com a evidencia mais forte de cada pagina elimina esse lixo.

2) NOMES DOS DIRETORIOS vem da leitura pagina a pagina (dict NOMES abaixo),
   nao do titulo do site — senao 10 sindicos ficariam todos chamados
   "Valorizando os sindicos de Santa Catarina - ASDESC".

Uso:  python _consolidar_sindicos.py            (so mostra o que vai gravar)
      python _consolidar_sindicos.py --gravar   (grava na planilha)
"""
import os, re, csv, sys, shutil
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font

BASE     = os.path.dirname(os.path.abspath(__file__))
COLETA   = os.path.join(BASE, "Comercial", "_coleta_sindicos.csv")
PLANILHA = os.path.join(BASE, "Comercial", "Sindicos_Profissionais_Brasil.xlsx")

LINHA_CABECALHO = 3
PRIMEIRA_LINHA  = 4

# Nome real lido na pagina do diretorio (telefone -> nome | cidade)
NOMES = {
    "5548984129560": ("Edgar Francis - B2R Condominios", "Florianopolis - SC"),
    "5547984541250": ("Guilherme Bora - Sindico Profissional", "Balneario Picarras - SC"),
    "5547984473915": ("Gustavo Camacho - Sindico Profissional", "Joinville - SC"),
    "5547984244510": ("Joice Honorio - Sindica Profissional", "Itajai/Balneario Camboriu - SC"),
    "5548984588008": ("Eduardo A. Patounas - Sindico Profissional", "Florianopolis - SC"),
    "5548999251684": ("Eliana Eidelwein - Habitar Multisindica", "Florianopolis - SC"),
    "5548999958888": ("Janaina Salvador - Sindica Profissional", "Florianopolis - SC"),
    "5548999893647": ("Luisa Candido Lopes - Sindica Profissional", "Florianopolis - SC"),
    "5547996251634": ("Rafael Bianchi Machado - Sindico Profissional", "Joinville - SC"),
    "5548998113863": ("Sindico Cipriano", "Biguacu/Sao Jose - SC"),
    "5548991467975": ("FOCO - Sindico Profissional", "Florianopolis - SC"),
}

# Paginas que sao diretorio de verdade: pode sair mais de um contato delas,
# porque cada numero foi conferido com o nome do dono na leitura manual.
DIRETORIOS = {
    "https://www.asdesc.com.br/",
    "https://condominiosc.com.br/guia-de-fornecedores/gestao/sindicos-profissionais",
}

UF_POR_DDD = {
    "11":"SP","12":"SP","13":"SP","14":"SP","15":"SP","16":"SP","17":"SP","18":"SP","19":"SP",
    "21":"RJ","22":"RJ","24":"RJ","27":"ES","28":"ES",
    "31":"MG","32":"MG","33":"MG","34":"MG","35":"MG","37":"MG","38":"MG",
    "41":"PR","42":"PR","43":"PR","44":"PR","45":"PR","46":"PR",
    "47":"SC","48":"SC","49":"SC","51":"RS","53":"RS","54":"RS","55":"RS",
    "61":"DF","62":"GO","63":"TO","64":"GO","65":"MT","66":"MT","67":"MS","68":"AC","69":"RO",
    "71":"BA","73":"BA","74":"BA","75":"BA","77":"BA","79":"SE",
    "81":"PE","82":"AL","83":"PB","84":"RN","85":"CE","86":"PI","87":"PE","88":"CE","89":"PI",
    "91":"PA","92":"AM","93":"PA","94":"PA","95":"RR","96":"AP","97":"AM","98":"MA","99":"MA",
}

FORCA = {"wa.me": 3, "tel:": 2, "texto": 1}


# Titulos que denunciam pagina de erro: o numero achado ali nao e da empresa.
LIXO = re.compile(r"(p[aá]gina n[aã]o encontrada|not found|404|erro|error)", re.I)

# Titulos genericos demais para servir de nome: usamos o dominio no lugar.
GENERICO = re.compile(r"^(s[ií]ndico profissional|s[ií]ndica profissional|home|in[ií]cio|"
                      r"contato|administra[cç][aã]o de condom[ií]nios|"
                      r"administradoras? de condom[ií]nios)$", re.I)


def dominio(url):
    m = re.match(r"https?://(?:www\.)?([^/]+)", url or "")
    return m.group(1) if m else ""


def consertar_acentos(t):
    """Titulo lido como latin-1 aparece como 'CondomÃ­nios'. Desfaz isso."""
    if "Ã" in t or "Â" in t:
        try:
            return t.encode("latin-1").decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            return t
    return t


def limpar_nome(titulo):
    """Titulo de pagina -> nome utilizavel."""
    t = consertar_acentos(re.sub(r"\s+", " ", titulo or "").strip())
    t = re.sub(r"^(Contato|Home|Início|Inicio|Quem somos)\s*[-–—|:]\s*", "", t, flags=re.I)
    for sep in ("|", "–", "—", "»", " - "):
        if sep in t:
            partes = [p.strip() for p in t.split(sep) if p.strip()]
            if partes:
                t = max(partes[:2], key=len)
            break
    return t[:60] or "(sem nome)"


def formatar(tel):
    """55DD9XXXXXXXX -> (DD) 9XXXX-XXXX"""
    n = tel[2:]
    return f"({n[:2]}) {n[2:7]}-{n[7:]}"


def main():
    gravar = "--gravar" in sys.argv

    rows = list(csv.DictReader(open(COLETA, encoding="utf-8")))

    # 1) melhor evidencia por pagina (diretorio conferido escapa da regra)
    melhor, soltos = {}, []
    for r in rows:
        if r["fonte"] in DIRETORIOS or r["telefone"] in NOMES:
            soltos.append(r)
            continue
        atual = melhor.get(r["fonte"])
        if atual is None or FORCA[r["evidencia"]] > FORCA[atual["evidencia"]]:
            melhor[r["fonte"]] = r
    candidatos = list(melhor.values()) + soltos

    # 2) telefone unico
    vistos, finais = set(), []
    for r in sorted(candidatos, key=lambda x: -FORCA[x["evidencia"]]):
        if r["telefone"] in vistos:
            continue
        if LIXO.search(r["nome"] or ""):
            continue            # numero achado em pagina de erro nao vale nada
        vistos.add(r["telefone"])
        nome, cidade = NOMES.get(r["telefone"], (limpar_nome(r["nome"]), ""))
        if GENERICO.match(nome):
            nome = f"{nome} ({dominio(r['fonte'])})"
        if not cidade:
            cidade = UF_POR_DDD.get(r["telefone"][2:4], "")
        finais.append({"nome": nome, "tel": r["telefone"], "cidade": cidade,
                       "evidencia": r["evidencia"], "fonte": r["fonte"]})

    # 3) tira quem ja esta na planilha
    wb = load_workbook(PLANILHA)
    ws = wb.worksheets[0]
    ja = set()
    for r in range(PRIMEIRA_LINHA, ws.max_row + 1):
        t = re.sub(r"\D", "", str(ws.cell(row=r, column=3).value or ""))
        if len(t) == 11:
            ja.add("55" + t)
    novos = [f for f in finais if f["tel"] not in ja]

    print(f"Coletados no CSV      : {len(rows)}")
    print(f"Depois de 1 por pagina: {len(finais)}")
    print(f"Ja na planilha        : {len(finais) - len(novos)}")
    print(f"NOVOS a incluir       : {len(novos)}\n")
    for i, f in enumerate(novos, 1):
        print(f"{i:3d}. {formatar(f['tel'])}  {f['evidencia']:6s}  {f['cidade']:32s} {f['nome'][:45]}")

    if not gravar:
        print("\n(nada gravado — rode com --gravar para incluir na planilha)")
        return

    backup = PLANILHA.replace(".xlsx", f"_bkp_{datetime.now():%Y%m%d-%H%M%S}.xlsx")
    shutil.copy2(PLANILHA, backup)
    print(f"\nCopia de seguranca: {backup}")

    linha = ws.max_row + 1
    seq = max((ws.cell(row=r, column=1).value or 0)
              for r in range(PRIMEIRA_LINHA, ws.max_row + 1)
              if isinstance(ws.cell(row=r, column=1).value, int))
    for f in novos:
        seq += 1
        ws.cell(row=linha, column=1, value=seq)
        ws.cell(row=linha, column=2, value=f["nome"])
        ws.cell(row=linha, column=3, value=formatar(f["tel"]))
        ws.cell(row=linha, column=4, value=f["cidade"])
        ws.cell(row=linha, column=5, value=f"Celular/WhatsApp ({f['evidencia']})")
        ws.cell(row=linha, column=6, value=f["fonte"])
        linha += 1

    wb.save(PLANILHA)
    print(f"Incluidos {len(novos)} contatos. Planilha: {PLANILHA}")


if __name__ == "__main__":
    main()
