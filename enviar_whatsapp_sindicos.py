# -*- coding: utf-8 -*-
"""
Envia a mensagem de apresentacao do Folha10-Simples para os sindicos
profissionais da planilha Comercial\\Sindicos_Profissionais_Brasil.xlsx,
com intervalo entre os envios para reduzir risco de bloqueio no WhatsApp.

Uso:
    python enviar_whatsapp_sindicos.py --teste          # SIMULA (nao envia)
    python enviar_whatsapp_sindicos.py                  # ENVIA de verdade
    python enviar_whatsapp_sindicos.py --intervalo 180  # segundos (padrao 180 = 3 min)
    python enviar_whatsapp_sindicos.py --intervalo 600 --intervalo-max 1200
    python enviar_whatsapp_sindicos.py --fonte coteibem # so os contatos daquela fonte

Com --intervalo-max a espera entre um envio e outro e sorteada dentro da faixa
[--intervalo, --intervalo-max] a cada envio, em vez de ser sempre a mesma. Um
ritmo cravado no relogio e o padrao mais facil de reconhecer como robo.

Janela de trabalho: por padrao so envia de 2a a 6a, das 9h as 17h. Fora disso o
script dorme ate a proxima abertura em vez de encerrar — da para deixar rodando
por dias que ele mesmo respeita o horario comercial. Ajuste com --hora-inicio /
--hora-fim, e --qualquer-dia para incluir sabado e domingo.

Diferencas para o enviar_whatsapp_clientes.py:
- A planilha tem 2 linhas de apresentacao antes do cabecalho: os dados
  comecam na linha 4, colunas: #, Nome, Telefone, Cidade, Tipo, Fonte.
- Telefone FIXO e PULADO, nao "corrigido". O normalizar_telefone() do outro
  script insere um 9 em numero de 10 digitos: o fixo (81) 3327-9227 viraria
  5581933279227, que e um celular de OUTRA pessoa.
- Log proprio (log_envio_sindicos.csv), separado da campanha de clientes.
"""
import os, re, sys, csv, time, random, argparse
from datetime import datetime

# Nome de contato vindo da web traz caractere fora do cp1252 (ex.: U+200E, a
# marca de direcao do texto). Sem isso o print estoura com UnicodeEncodeError e
# derruba a campanha inteira no meio.
for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

import requests
import urllib3
from openpyxl import load_workbook
from dotenv import load_dotenv

BASE     = os.path.dirname(os.path.abspath(__file__))
PLANILHA = os.path.join(BASE, "Comercial", "Sindicos_Profissionais_Brasil.xlsx")
MSG_FILE = os.path.join(BASE, "mensagem_whatsapp_apresentacao.txt")
LOG_CSV  = os.path.join(BASE, "log_envio_sindicos.csv")
LINHA_CABECALHO = 3         # linhas 1-2 = apresentacao
PRIMEIRA_LINHA  = 4

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
load_dotenv(os.path.join(BASE, ".env"))
ZAPI_INSTANCE     = os.getenv("ZAPI_INSTANCE", "").strip()
ZAPI_TOKEN        = os.getenv("ZAPI_TOKEN", "").strip()
ZAPI_CLIENT_TOKEN = os.getenv("ZAPI_CLIENT_TOKEN", "").strip()
ZAPI_SEND_URL   = f"https://api.z-api.io/instances/{ZAPI_INSTANCE}/token/{ZAPI_TOKEN}/send-text"
ZAPI_STATUS_URL = f"https://api.z-api.io/instances/{ZAPI_INSTANCE}/token/{ZAPI_TOKEN}/status"
HEADERS = {
    "Client-Token": ZAPI_CLIENT_TOKEN,
    "Content-Type": "application/json",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
}


def so_numeros(t):
    return re.sub(r"\D", "", t or "")


def classificar(tel_bruto):
    """Retorna (telefone_normalizado, motivo_de_pular).
    Celular BR = 55 + DDD(2) + 9 + 8 digitos = 13 chars."""
    t = so_numeros(tel_bruto)
    if t.startswith("55") and len(t) in (12, 13):
        t = t[2:]
    if len(t) == 10:
        return "", "fixo (nao inserimos o 9 — viraria o celular de outra pessoa)"
    if len(t) != 11:
        return "", f"formato inesperado ({len(t)} digitos)"
    if t[2] != "9":
        return "", "nao parece celular (nao comeca com 9 apos o DDD)"
    return "55" + t, ""


def _req(metodo, url, **kw):
    """requests com fallback: se a verificacao SSL falhar no ambiente local
    (proxy/antivirus interceptando), refaz a chamada sem verificar."""
    try:
        return requests.request(metodo, url, **kw)
    except requests.exceptions.SSLError:
        return requests.request(metodo, url, verify=False, **kw)


def zapi_conectada():
    if not (ZAPI_INSTANCE and ZAPI_TOKEN and ZAPI_CLIENT_TOKEN):
        return False, "Z-API nao configurada no .env"
    try:
        r = _req("GET", ZAPI_STATUS_URL, headers={"Client-Token": ZAPI_CLIENT_TOKEN},
                 timeout=20, verify=False)
        if r.status_code != 200:
            return False, f"HTTP {r.status_code}: {r.text[:200]}"
        info = r.json()
        if info.get("connected") is True:
            return True, "conectada"
        return False, f"desconectada: {info}"
    except Exception as e:
        return False, str(e)


def enviar(telefone, mensagem):
    try:
        r = _req("POST", ZAPI_SEND_URL, headers=HEADERS,
                 json={"phone": telefone, "message": mensagem}, timeout=30, verify=False)
        if r.status_code in (200, 201):
            return True, r.text[:200]
        return False, f"HTTP {r.status_code}: {r.text[:200]}"
    except Exception as e:
        return False, str(e)


def ja_enviados_ok():
    feitos = set()
    if os.path.exists(LOG_CSV):
        with open(LOG_CSV, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                if row.get("status") == "OK":
                    feitos.add(row.get("telefone"))
    return feitos


# Handle do mutex em variavel de modulo: se ele for coletado, o Windows libera
# a trava e uma segunda campanha entra.
_MUTEX = None


def instancia_unica():
    """True se esta e a unica campanha rodando; False se ja existe outra.

    O Agendador dispara todo dia as 9h, mas o script nao encerra quando a janela
    fecha as 17h — dorme ate a proxima abertura. Sem esta trava, cada dia util
    somava mais uma instancia viva percorrendo a MESMA lista, e o contato recebia
    uma mensagem de cada uma (em 12/08/2026 havia 4 processos: 07, 10, 11 e 12/08).
    Mutex nomeado do Windows porque o sistema o libera sozinho quando o processo
    morre — nao fica trava velha para trombar depois de um encerramento feio.
    """
    global _MUTEX
    if os.name != "nt":
        return True
    import ctypes
    ERROR_ALREADY_EXISTS = 183
    k32 = ctypes.windll.kernel32
    _MUTEX = k32.CreateMutexW(None, False, "Folha10_CampanhaSindicos")
    return k32.GetLastError() != ERROR_ALREADY_EXISTS


def carregar_contatos():
    wb = load_workbook(PLANILHA, read_only=True)
    ws = wb.active
    # A coluna DataUltimaMensagem e preenchida pelo _marcar_planilha_sindicos.py
    # e serve de segunda trava contra enviar duas vezes para o mesmo contato:
    # quem ja tem data ali nao entra na fila, mesmo que o log CSV se perca.
    cab = next(ws.iter_rows(min_row=LINHA_CABECALHO, max_row=LINHA_CABECALHO,
                            values_only=True), ())
    i_data = next((i for i, v in enumerate(cab)
                   if str(v or "").strip() == "DataUltimaMensagem"), None)

    contatos = []
    for row in ws.iter_rows(min_row=PRIMEIRA_LINHA, values_only=True):
        nome  = (row[1] or "").strip() if row[1] else ""
        tel   = (row[2] or "").strip() if row[2] else ""
        cid   = (row[3] or "").strip() if len(row) > 3 and row[3] else ""
        fonte = (row[5] or "").strip() if len(row) > 5 and row[5] else ""
        if not nome or not tel:
            continue
        ja_msg = ""
        if i_data is not None and len(row) > i_data and row[i_data]:
            ja_msg = str(row[i_data]).strip()
        tel_norm, motivo = classificar(tel)
        contatos.append({"nome": nome, "tel_orig": tel, "cidade": cid,
                         "fonte": fonte, "tel": tel_norm, "pular": motivo,
                         "ja_msg": ja_msg})
    wb.close()
    return contatos


def marcar_planilha():
    """Grava DataUltimaMensagem/UltimaMensagem na planilha a partir do log.

    Roda ao fechar a janela do dia e no fim da campanha. Se a planilha estiver
    aberta no Excel a gravacao falha — nao e fatal, o log CSV continua sendo a
    trava real contra envio repetido e a proxima marcacao pega o atrasado.
    """
    try:
        import _marcar_planilha_sindicos as M
        print("      marcando a planilha com quem ja recebeu...", flush=True)
        M.main(verbose=False)
    except Exception as e:
        print(f"      [AVISO] nao deu para marcar a planilha agora: {e}", flush=True)
        print("      (o log_envio_sindicos.csv continua valendo como controle)", flush=True)


# ── Janela de trabalho (dia da semana + faixa de horario) ──────────────────
def dentro_da_janela(dt, h_ini, h_fim, dias_uteis):
    if dias_uteis and dt.weekday() >= 5:        # 5=sabado, 6=domingo
        return False
    return h_ini <= dt.hour < h_fim


def proxima_abertura(dt, h_ini, h_fim, dias_uteis):
    """Primeiro instante >= dt que cai dentro da janela."""
    from datetime import timedelta
    c = dt
    for _ in range(21):                          # no maximo 3 semanas a frente
        if (not dias_uteis) or c.weekday() < 5:
            if c.hour < h_ini:
                return c.replace(hour=h_ini, minute=0, second=0, microsecond=0)
            if c.hour < h_fim:
                return c
        c = (c + timedelta(days=1)).replace(hour=h_ini, minute=0, second=0,
                                            microsecond=0)
    return c


def aguardar_janela(h_ini, h_fim, dias_uteis, teste=False, ao_fechar=None):
    """Dorme ate a janela abrir. Em --teste so avisa e segue."""
    fechou = False
    while True:
        agora = datetime.now()
        if dentro_da_janela(agora, h_ini, h_fim, dias_uteis):
            return
        alvo = proxima_abertura(agora, h_ini, h_fim, dias_uteis)
        if teste:
            print(f"      [TESTE] fora da janela — retomaria em "
                  f"{alvo:%d/%m/%Y %H:%M}", flush=True)
            return
        # A janela acabou de fechar: consolida o dia na planilha antes de dormir.
        if ao_fechar and not fechou:
            ao_fechar()
            fechou = True
        seg = max(1, int((alvo - agora).total_seconds()))
        print(f"      fora da janela ({agora:%d/%m %H:%M}) — retoma "
              f"{alvo:%a %d/%m %H:%M} (em {seg/3600:.1f} h)", flush=True)
        while seg > 0:                           # dorme em fatias de 15 min
            fatia = min(seg, 900)
            time.sleep(fatia)
            seg -= fatia


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--teste", action="store_true", help="Simula, nao envia")
    ap.add_argument("--limite", type=int, default=0, help="Envia so aos N primeiros")
    ap.add_argument("--intervalo", type=int, default=180, help="Segundos entre envios (padrao 180 = 3 min)")
    ap.add_argument("--intervalo-max", type=int, default=0, dest="intervalo_max",
                    help="Se informado, a espera e sorteada entre --intervalo e este valor "
                         "a cada envio (ex.: --intervalo 600 --intervalo-max 1200 = 10 a 20 min)")
    ap.add_argument("--espera-inicial", type=int, default=0, dest="espera_inicial",
                    help="Segundos de espera ANTES do primeiro envio. Serve para retomar "
                         "uma campanha sem encurtar o intervalo em relacao ao ultimo envio.")
    ap.add_argument("--fonte", default="", help="So os contatos cuja Fonte contenha esse texto")
    ap.add_argument("--hora-inicio", type=int, default=9, dest="hora_inicio",
                    help="Hora em que a janela abre (padrao 9)")
    ap.add_argument("--hora-fim", type=int, default=17, dest="hora_fim",
                    help="Hora em que a janela fecha (padrao 17 — o ultimo envio sai antes das 17h)")
    ap.add_argument("--qualquer-dia", action="store_true", dest="qualquer_dia",
                    help="Envia tambem no sabado e no domingo (padrao: so 2a a 6a)")
    ap.add_argument("--ignorar-trava", action="store_true", dest="ignorar_trava",
                    help="Roda mesmo com outra campanha aberta (so para depurar)")
    args = ap.parse_args()
    dias_uteis = not args.qualquer_dia

    if not args.teste and not args.ignorar_trava and not instancia_unica():
        print("[ABORTADO] ja existe uma campanha rodando nesta maquina.", flush=True)
        print("Duas campanhas ao mesmo tempo mandam a mesma mensagem duas vezes "
              "para o mesmo contato.", flush=True)
        sys.exit(0)

    # Faixa da espera. Sem --intervalo-max (ou com valor menor que o piso) o
    # comportamento e o de antes: espera fixa igual a --intervalo.
    esp_min = max(1, args.intervalo)
    esp_max = max(esp_min, args.intervalo_max or 0)
    esp_medio = (esp_min + esp_max) / 2

    with open(MSG_FILE, encoding="utf-8") as f:
        mensagem = f.read().strip()

    contatos = carregar_contatos()
    if args.fonte:
        alvo_fonte = args.fonte.lower()
        contatos = [c for c in contatos if alvo_fonte in (c["fonte"] or "").lower()]
    validos  = [c for c in contatos if not c["pular"]]
    pulados_fmt = [c for c in contatos if c["pular"]]
    if args.limite:
        validos = validos[: args.limite]

    feitos = ja_enviados_ok()
    faltam = len([c for c in validos if c["tel"] not in feitos and not c["ja_msg"]])
    marcados_planilha = len([c for c in validos if c["ja_msg"]])

    # quantos cabem por dia dentro da janela, e em quantos dias isso termina
    por_dia = max(1, int((args.hora_fim - args.hora_inicio) * 3600 / esp_medio))
    dias    = (faltam + por_dia - 1) // por_dia

    print("=" * 70, flush=True)
    print(f"Planilha  : {PLANILHA}", flush=True)
    if args.fonte:
        print(f"Filtro    : fonte contem '{args.fonte}'", flush=True)
    print(f"Contatos  : {len(contatos)}  |  enviaveis: {len(validos)}  |  fora: {len(pulados_fmt)}", flush=True)
    for c in pulados_fmt:
        print(f"   FORA: {c['nome']} ({c['tel_orig']}) — {c['pular']}", flush=True)
    print(f"Ja OK     : {len(feitos)} no log (os desta lista serao pulados)", flush=True)
    print(f"Ja marcado: {marcados_planilha} com data na planilha (tambem pulados)", flush=True)
    print(f"A enviar  : {faltam}", flush=True)
    if esp_max > esp_min:
        print(f"Intervalo : aleatorio de {esp_min}s a {esp_max}s "
              f"({esp_min/60:.1f} a {esp_max/60:.1f} min, media {esp_medio/60:.1f})", flush=True)
    else:
        print(f"Intervalo : {esp_min}s ({esp_min/60:.1f} min) fixo", flush=True)
    print(f"Janela    : {args.hora_inicio}h as {args.hora_fim}h, "
          f"{'2a a 6a' if dias_uteis else 'todos os dias'}", flush=True)
    print(f"Ritmo     : ate {por_dia}/dia  ->  ~{dias} dia(s) de janela", flush=True)
    print(f"Modo      : {'TESTE (nao envia)' if args.teste else 'ENVIO REAL'}", flush=True)
    print("=" * 70, flush=True)

    if not args.teste:
        ok, msg = zapi_conectada()
        if not ok:
            print(f"[ABORTADO] WhatsApp/Z-API nao conectada: {msg}", flush=True)
            print("Reconecte em https://app.z-api.io e rode de novo.", flush=True)
            sys.exit(1)
        print(f"[Z-API] {msg}", flush=True)

    novo_log = not os.path.exists(LOG_CSV)
    logf = open(LOG_CSV, "a", newline="", encoding="utf-8")
    logw = csv.writer(logf)
    if novo_log:
        logw.writerow(["datahora", "nome", "telefone", "status", "detalhe"])

    total = len(validos)
    enviados = pulados = falhas = 0
    esperou_inicial = False

    for i, c in enumerate(validos, 1):
        tel = c["tel"]
        prefixo = f"[{i}/{total}] {c['nome']} -> {tel}"

        if tel in feitos:
            print(f"{prefixo}  ::  PULADO (ja enviado antes)", flush=True)
            pulados += 1
            continue
        if c["ja_msg"]:
            print(f"{prefixo}  ::  PULADO (planilha ja marca envio em {c['ja_msg']})", flush=True)
            pulados += 1
            continue

        # Fora do horario de trabalho o script dorme ate a janela reabrir.
        aguardar_janela(args.hora_inicio, args.hora_fim, dias_uteis, args.teste,
                        ao_fechar=(None if args.teste else marcar_planilha))

        # Espera so antes do 1o envio de verdade — depois dos pulados, para
        # nao encurtar o intervalo em relacao ao ultimo envio da rodada anterior.
        if args.espera_inicial and not esperou_inicial and not args.teste:
            print(f"      espera inicial de {args.espera_inicial}s "
                  f"({args.espera_inicial/60:.1f} min) antes de retomar...", flush=True)
            time.sleep(args.espera_inicial)
            esperou_inicial = True

        # Relê o log agora, na hora do envio: entre a largada e este ponto podem
        # ter se passado horas (ou dias, se dormiu fora da janela) e outra
        # campanha pode ter atendido este contato. O `feitos` da largada e uma
        # foto do passado. Arquivo pequeno — o custo e irrelevante perto dos
        # 10 a 20 min de espera entre um envio e outro.
        if not args.teste and tel in ja_enviados_ok():
            print(f"{prefixo}  ::  PULADO (ja consta no log — outra campanha enviou)",
                  flush=True)
            feitos.add(tel)
            pulados += 1
            continue

        agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if args.teste:
            print(f"{prefixo}  ::  [TESTE] enviaria ({len(mensagem)} chars)", flush=True)
            logw.writerow([agora, c["nome"], tel, "TESTE", "simulado"])
            enviados += 1
        else:
            ok, detalhe = enviar(tel, mensagem)
            status = "OK" if ok else "FALHA"
            print(f"{agora}  {prefixo}  ::  {status}  {'' if ok else detalhe}", flush=True)
            logw.writerow([agora, c["nome"], tel, status, detalhe])
            if ok:
                feitos.add(tel)
                enviados += 1
            else:
                falhas += 1
        logf.flush()

        if i < total:
            espera = 2 if args.teste else random.randint(esp_min, esp_max)
            print(f"      aguardando {espera}s ({espera/60:.1f} min)...", flush=True)
            time.sleep(espera)

    logf.close()
    if not args.teste and enviados:
        marcar_planilha()
    print("=" * 70, flush=True)
    print(f"FIM. Enviados OK: {enviados} | Falhas: {falhas} | Pulados: {pulados}", flush=True)
    print(f"Log: {LOG_CSV}", flush=True)


if __name__ == "__main__":
    main()
