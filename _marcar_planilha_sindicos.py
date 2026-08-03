# -*- coding: utf-8 -*-
"""
Marca na planilha Comercial\\Sindicos_Profissionais_Brasil.xlsx quem ja
recebeu mensagem, lendo o log_envio_sindicos.csv (so as linhas com status OK).

- Cria (se ainda nao existirem) as colunas DataUltimaMensagem e UltimaMensagem.
- Cria (se ainda nao existir) a aba "Mensagens" com o codigo e o texto de cada
  mensagem enviada, para saber depois o que cada contato recebeu.
- Faz copia de seguranca da planilha antes de gravar.

Pode rodar quantas vezes quiser: so preenche quem esta como OK no log, entao
rodar de novo no fim da campanha completa os que faltavam.

Uso:  python _marcar_planilha_sindicos.py
"""
import os, re, sys, csv, shutil
from datetime import datetime

# Nome de contato com caractere fora do cp1252 (ex.: U+200E) estourava o print.
for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

BASE     = os.path.dirname(os.path.abspath(__file__))
PLANILHA = os.path.join(BASE, "Comercial", "Sindicos_Profissionais_Brasil.xlsx")
LOG_CSV  = os.path.join(BASE, "log_envio_sindicos.csv")
MSG_FILE = os.path.join(BASE, "mensagem_whatsapp_apresentacao.txt")

LINHA_CABECALHO = 3
PRIMEIRA_LINHA  = 4
COL_TELEFONE    = 3          # C

CODIGO_MSG = "Mensagem 001"
ASSUNTO_MSG = "Apresentacao do Folha10-Simples - 1o mes gratuito"


def so_numeros(t):
    return re.sub(r"\D", "", str(t or ""))


def normalizar(tel):
    """Mesma regra do enviar_whatsapp_sindicos.py, so para casar com o log."""
    t = so_numeros(tel)
    if t.startswith("55") and len(t) in (12, 13):
        t = t[2:]
    return "55" + t if len(t) == 11 else ""


def enviados_ok():
    """{telefone_normalizado: datahora do envio} das linhas com status OK."""
    out = {}
    if not os.path.exists(LOG_CSV):
        return out
    with open(LOG_CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row.get("status") == "OK":
                out[row["telefone"]] = row["datahora"]
    return out


def achar_ou_criar_coluna(ws, titulo):
    """Devolve o indice da coluna com esse titulo, criando no fim se faltar."""
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(row=LINHA_CABECALHO, column=col).value or "").strip() == titulo:
            return col
    col = ws.max_column + 1
    c = ws.cell(row=LINHA_CABECALHO, column=col, value=titulo)
    c.font = Font(bold=True)
    ws.column_dimensions[c.column_letter].width = max(20, len(titulo) + 4)
    return col


def aba_mensagens(wb):
    """Cria a aba Mensagens com o texto de cada codigo de mensagem."""
    if "Mensagens" in wb.sheetnames:
        ws = wb["Mensagens"]
    else:
        ws = wb.create_sheet("Mensagens")
        ws["A1"] = "Mensagens enviadas - historico dos textos"
        ws["A1"].font = Font(bold=True, size=12)
        cabecalhos = ["Codigo", "Data", "Assunto", "Texto enviado"]
        for i, h in enumerate(cabecalhos, start=1):
            c = ws.cell(row=2, column=i, value=h)
            c.font = Font(bold=True)
        for letra, larg in (("A", 14), ("B", 12), ("C", 46), ("D", 100)):
            ws.column_dimensions[letra].width = larg

    # Ja existe a linha desse codigo?
    for r in range(3, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").strip() == CODIGO_MSG:
            return ws

    with open(MSG_FILE, encoding="utf-8") as f:
        texto = f.read().strip()

    linha = max(3, ws.max_row + 1)
    ws.cell(row=linha, column=1, value=CODIGO_MSG).font = Font(bold=True)
    ws.cell(row=linha, column=2, value=datetime.now().strftime("%d/%m/%Y"))
    ws.cell(row=linha, column=3, value=ASSUNTO_MSG)
    cel = ws.cell(row=linha, column=4, value=texto)
    cel.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[linha].height = 90
    return ws


def main(verbose=True):
    ok = enviados_ok()
    print(f"Log: {len(ok)} envio(s) com status OK")

    backup = PLANILHA.replace(".xlsx", f"_bkp_{datetime.now():%Y%m%d-%H%M%S}.xlsx")
    shutil.copy2(PLANILHA, backup)
    print(f"Copia de seguranca: {backup}")

    wb = load_workbook(PLANILHA)
    ws = wb.worksheets[0]

    col_data = achar_ou_criar_coluna(ws, "DataUltimaMensagem")
    col_msg  = achar_ou_criar_coluna(ws, "UltimaMensagem")

    marcados = pendentes = 0
    for r in range(PRIMEIRA_LINHA, ws.max_row + 1):
        tel_bruto = ws.cell(row=r, column=COL_TELEFONE).value
        if not tel_bruto:
            continue
        tel = normalizar(tel_bruto)
        nome = ws.cell(row=r, column=2).value
        if tel and tel in ok:
            data_envio = datetime.strptime(ok[tel], "%Y-%m-%d %H:%M:%S")
            ws.cell(row=r, column=col_data, value=data_envio.strftime("%d/%m/%Y"))
            ws.cell(row=r, column=col_msg,  value=CODIGO_MSG)
            marcados += 1
        else:
            pendentes += 1
            if verbose:
                print(f"   sem marcar: {nome} ({tel_bruto})")

    aba_mensagens(wb)
    wb.save(PLANILHA)

    print(f"Marcados: {marcados} | Sem marcar: {pendentes}")
    print(f"Planilha gravada: {PLANILHA}")


if __name__ == "__main__":
    main()
