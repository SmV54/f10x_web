# -*- coding: utf-8 -*-
"""Inclui na planilha a leva de PEQUENAS EMPRESAS (tipo de contato diferente).

Entrada: Comercial\\_todas_fichas.json — varredura de todas as ~15 mil fichas do
diretorio (rodar antes: python _adm_2_fichas.py --todas).

Quem e do ramo condominial NAO entra aqui: ja esta na planilha como
"Sindico/Administradora". O resto sao fornecedores de condominio (portaria,
limpeza, conservacao, manutencao...), que e o outro publico da mensagem:
pequena empresa com folha de pagamento.

Ordem de entrada: primeiro as categorias de mao de obra intensiva, que sao as
que mais tem funcionario e mais perto estao do dia a dia do condominio.

Mesmas regras de qualidade das outras etapas: so celular, telefone e nome de
empresa nao repetidos, e o numero tem que ter vindo do cadastro da empresa.

Uso:  python _adm_7_pequenas.py [--limite N]        (mostra o que entraria)
      python _adm_7_pequenas.py --limite N --gravar (grava)
"""
import os, re, sys, json, shutil, unicodedata
from datetime import datetime

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

BASE     = os.path.dirname(os.path.abspath(__file__))
TODAS    = os.path.join(BASE, "Comercial", "_todas_fichas.json")
PLANILHA = os.path.join(BASE, "Comercial", "Sindicos_Profissionais_Brasil.xlsx")

PRIMEIRA_LINHA = 4
COL_TIPO       = 9
TIPO           = "Pequena empresa"

sys.path.insert(0, BASE)
from _adm_3_planilha import (normalizar, formatar, limpar_nome, chave_nome,
                             dominio, UF_POR_DDD, LIXO, CATEGORIAS_OK, _chave_cat)

# Categorias de mao de obra intensiva entram primeiro
PRIORIDADE = [
    "portaria", "seguranca", "vigilancia", "limpeza", "conservacao", "zelador",
    "jardinagem", "manutencao", "obras", "pintura", "elevador", "piscina",
    "dedetiza", "controle de pragas", "reforma", "hidraulic", "eletric",
]


def _sem_acento(t):
    t = unicodedata.normalize("NFKD", str(t or ""))
    return "".join(c for c in t if not unicodedata.combining(c)).lower()


def peso(cats):
    txt = _sem_acento(" ".join(cats or []))
    for i, p in enumerate(PRIORIDADE):
        if p in txt:
            return i
    return len(PRIORIDADE)


def main():
    gravar = "--gravar" in sys.argv
    limite = 0
    for a in sys.argv:
        if a.startswith("--limite"):
            try:
                limite = int(a.split("=")[1]) if "=" in a else int(sys.argv[sys.argv.index(a) + 1])
            except Exception:
                limite = 0

    if not os.path.exists(TODAS):
        print("Rode antes: python _adm_2_fichas.py --todas")
        return

    fichas = json.load(open(TODAS, encoding="utf-8"))
    print(f"fichas na varredura geral: {len(fichas)}")

    cand = []
    for f in fichas.values():
        cats = f.get("categorias") or []
        # do ramo condominial ja entrou como Sindico/Administradora
        if any(_chave_cat(c) in CATEGORIAS_OK for c in cats):
            continue
        tel = normalizar(f.get("telefone"))
        if not tel:
            continue
        nome = limpar_nome(f.get("nome"))
        if not nome or LIXO.search(nome):
            continue
        cidade = f.get("cidade") or ""
        uf = f.get("uf") or UF_POR_DDD.get(tel[2:4], "")
        cand.append({
            "nome": nome, "tel": tel,
            "cidade": f"{cidade} - {uf}".strip(" -"),
            "ramo": (cats[0] if cats else ""),
            "peso": peso(cats),
            "fonte": f.get("fonte", ""),
        })

    # telefone unico dentro da propria leva
    vistos, unicos = set(), []
    for c in sorted(cand, key=lambda x: (x["peso"], x["nome"])):
        if c["tel"] in vistos:
            continue
        vistos.add(c["tel"])
        unicos.append(c)

    wb = load_workbook(PLANILHA)
    ws = wb.worksheets[0]
    ja, ja_nome = set(), set()
    for r in range(PRIMEIRA_LINHA, ws.max_row + 1):
        t = re.sub(r"\D", "", str(ws.cell(row=r, column=3).value or ""))
        if len(t) == 11:
            ja.add("55" + t)
        n = chave_nome(ws.cell(row=r, column=2).value)
        if n:
            ja_nome.add(n)

    novos = []
    for c in unicos:
        if c["tel"] in ja:
            continue
        n = chave_nome(c["nome"])
        if n and n in ja_nome:
            continue
        if n:
            ja_nome.add(n)
        novos.append(c)
        if limite and len(novos) >= limite:
            break

    print(f"candidatas (fora do ramo condominial, com celular): {len(unicos)}")
    print(f"NOVAS a incluir: {len(novos)}" + (f"  (limite {limite})" if limite else ""))
    porramo = {}
    for n in novos:
        porramo[n["ramo"]] = porramo.get(n["ramo"], 0) + 1
    for k, v in sorted(porramo.items(), key=lambda x: -x[1])[:15]:
        print(f"   {v:4d}  {k}")

    for i, c in enumerate(novos[:15], 1):
        print(f"{i:4d}. {formatar(c['tel'])}  {c['cidade'][:26]:26s} {c['nome'][:40]:40s} {c['ramo'][:28]}")

    if not gravar:
        print("\n(nada gravado — rode com --gravar)")
        return

    bkp = PLANILHA.replace(".xlsx", f"_bkp_{datetime.now():%Y%m%d-%H%M%S}.xlsx")
    shutil.copy2(PLANILHA, bkp)
    print("\nCopia de seguranca:", bkp)

    linha = ws.max_row + 1
    seq = max((ws.cell(row=r, column=1).value or 0)
              for r in range(PRIMEIRA_LINHA, ws.max_row + 1)
              if isinstance(ws.cell(row=r, column=1).value, int))
    for c in novos:
        seq += 1
        ws.cell(row=linha, column=1, value=seq)
        ws.cell(row=linha, column=2, value=c["nome"])
        ws.cell(row=linha, column=3, value=formatar(c["tel"]))
        ws.cell(row=linha, column=4, value=c["cidade"])
        ws.cell(row=linha, column=5, value="Celular/WhatsApp (ficha)")
        ws.cell(row=linha, column=6, value=c["fonte"])
        ws.cell(row=linha, column=COL_TIPO, value=TIPO)
        linha += 1

    wb.save(PLANILHA)
    print(f"Incluidos {len(novos)} contatos como {TIPO!r}.")


if __name__ == "__main__":
    main()
