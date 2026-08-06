# -*- coding: utf-8 -*-
"""
Mesmo controle do _marcar_planilha_sindicos.py, agora na planilha
Clientes_Cadastrados_Folha10.xlsx, lendo o log_envio_whatsapp.csv.

- Cria as colunas DataUltimaMensagem e UltimaMensagem (so as linhas com OK).
- Cria a aba "Mensagens" com o texto da campanha de 18/07/2026 (Mensagem 000).
- Copia de seguranca antes de gravar. Pode rodar de novo sem duplicar nada.

ATENCAO ao casamento dos telefones: aquele disparo usou o
normalizar_telefone() do enviar_whatsapp_clientes.py, que INSERE um 9 em
numero de 10 digitos. Para achar a linha certa no log e preciso repetir a
mesma regra — e as linhas em que isso aconteceu ficam sinalizadas na saida,
porque a mensagem pode ter ido para o celular de outra pessoa.

Uso:  python _marcar_planilha_clientes.py
"""
import os, re, csv, shutil
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

BASE     = os.path.dirname(os.path.abspath(__file__))
PLANILHA = os.path.join(BASE, "Clientes_Cadastrados_Folha10.xlsx")
LOG_CSV  = os.path.join(BASE, "log_envio_whatsapp.csv")
MSG_FILE = os.path.join(BASE, "mensagem_whatsapp.txt")

LINHA_CABECALHO = 1
PRIMEIRA_LINHA  = 2
COL_NOME        = 2          # B
COL_TELEFONE    = 4          # D

CODIGO_MSG  = "Mensagem 000"
ASSUNTO_MSG = "Campanha de reativacao - Folha10-Simples na web (julho/2026)"


def so_numeros(t):
    return re.sub(r"\D", "", str(t or ""))


def normalizar_como_no_envio(t):
    """Copia fiel do normalizar_telefone() do enviar_whatsapp_clientes.py.
    Retorna (telefone, inseriu_9)."""
    t = so_numeros(t)
    if t.startswith("55") and len(t) in (12, 13):
        t = t[2:]
    inseriu = False
    if len(t) == 10:
        t = t[:2] + "9" + t[2:]
        inseriu = True
    return "55" + t, inseriu


def enviados_ok():
    out = {}
    if not os.path.exists(LOG_CSV):
        return out
    with open(LOG_CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row.get("status") == "OK":
                out[row["telefone"]] = row["datahora"]
    return out


def achar_ou_criar_coluna(ws, titulo):
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(row=LINHA_CABECALHO, column=col).value or "").strip() == titulo:
            return col
    col = ws.max_column + 1
    c = ws.cell(row=LINHA_CABECALHO, column=col, value=titulo)
    c.font = Font(bold=True)
    ws.column_dimensions[c.column_letter].width = max(20, len(titulo) + 4)
    return col


def aba_mensagens(wb, data_campanha):
    if "Mensagens" in wb.sheetnames:
        ws = wb["Mensagens"]
    else:
        ws = wb.create_sheet("Mensagens")
        ws["A1"] = "Mensagens enviadas - historico dos textos"
        ws["A1"].font = Font(bold=True, size=12)
        for i, h in enumerate(["Codigo", "Data", "Assunto", "Texto enviado"], start=1):
            ws.cell(row=2, column=i, value=h).font = Font(bold=True)
        for letra, larg in (("A", 14), ("B", 12), ("C", 46), ("D", 100)):
            ws.column_dimensions[letra].width = larg

    for r in range(3, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value or "").strip() == CODIGO_MSG:
            return ws

    with open(MSG_FILE, encoding="utf-8") as f:
        texto = f.read().strip()

    linha = max(3, ws.max_row + 1)
    ws.cell(row=linha, column=1, value=CODIGO_MSG).font = Font(bold=True)
    ws.cell(row=linha, column=2, value=data_campanha)
    ws.cell(row=linha, column=3, value=ASSUNTO_MSG)
    cel = ws.cell(row=linha, column=4, value=texto)
    cel.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[linha].height = 90
    return ws


def main():
    ok = enviados_ok()
    print(f"Log: {len(ok)} envio(s) com status OK")
    data_campanha = (min(ok.values())[:10] if ok else "")
    if data_campanha:
        d = datetime.strptime(data_campanha, "%Y-%m-%d")
        data_campanha = d.strftime("%d/%m/%Y")

    backup = PLANILHA.replace(".xlsx", f"_bkp_{datetime.now():%Y%m%d-%H%M%S}.xlsx")
    shutil.copy2(PLANILHA, backup)
    print(f"Copia de seguranca: {backup}")

    wb = load_workbook(PLANILHA)
    ws = wb.worksheets[0]

    col_data = achar_ou_criar_coluna(ws, "DataUltimaMensagem")
    col_msg  = achar_ou_criar_coluna(ws, "UltimaMensagem")

    marcados = []
    sem_marcar = []
    com_9_inserido = []
    for r in range(PRIMEIRA_LINHA, ws.max_row + 1):
        tel_bruto = ws.cell(row=r, column=COL_TELEFONE).value
        nome = ws.cell(row=r, column=COL_NOME).value
        if not tel_bruto:
            continue
        tel, inseriu = normalizar_como_no_envio(tel_bruto)
        if tel in ok:
            envio = datetime.strptime(ok[tel], "%Y-%m-%d %H:%M:%S")
            ws.cell(row=r, column=col_data, value=envio.strftime("%d/%m/%Y"))
            ws.cell(row=r, column=col_msg,  value=CODIGO_MSG)
            marcados.append(nome)
            if inseriu:
                com_9_inserido.append((nome, tel_bruto, tel))
        else:
            sem_marcar.append((nome, tel_bruto))

    aba_mensagens(wb, data_campanha)
    wb.save(PLANILHA)

    print(f"\nMarcados  : {len(marcados)}")
    print(f"Sem marcar: {len(sem_marcar)}")
    for nome, tel in sem_marcar:
        print(f"   {nome} ({tel})")
    if com_9_inserido:
        print(f"\nATENCAO - {len(com_9_inserido)} numero(s) tinham 8 digitos e o envio "
              f"inseriu um 9. Confira se o destino estava certo:")
        for nome, orig, virou in com_9_inserido:
            print(f"   {nome}: {orig} -> {virou}")
    print(f"\nPlanilha gravada: {PLANILHA}")


if __name__ == "__main__":
    main()
