# -*- coding: utf-8 -*-
"""
Envia UMA mensagem de WhatsApp (mesmo texto) para cada cliente da planilha,
com intervalo de 2 minutos entre cada envio, para reduzir risco de bloqueio Z-API.

Uso:
    python enviar_whatsapp_clientes.py --teste     # SIMULA (nao envia nada)
    python enviar_whatsapp_clientes.py             # ENVIA de verdade
    python enviar_whatsapp_clientes.py --limite 5  # envia so aos 5 primeiros
    python enviar_whatsapp_clientes.py --intervalo 120   # segundos (padrao 120)

- Le as credenciais Z-API do .env do projeto (mesmas do app.py).
- Le a mensagem do arquivo mensagem_whatsapp.txt ({nome} vira o 1o nome do cliente).
- Checa se a instancia Z-API esta conectada ANTES de comecar.
- Grava log em log_envio_whatsapp.csv. Se rodar de novo, PULA quem ja recebeu OK.
"""
import os, re, sys, csv, time, argparse
from datetime import datetime

import requests
from openpyxl import load_workbook
from dotenv import load_dotenv

BASE = os.path.dirname(os.path.abspath(__file__))
PLANILHA = os.path.join(BASE, "Clientes_Cadastrados_Folha10.xlsx")
MSG_FILE = os.path.join(BASE, "mensagem_whatsapp.txt")
LOG_CSV  = os.path.join(BASE, "log_envio_whatsapp.csv")

load_dotenv(os.path.join(BASE, ".env"))
ZAPI_INSTANCE     = os.getenv("ZAPI_INSTANCE", "").strip()
ZAPI_TOKEN        = os.getenv("ZAPI_TOKEN", "").strip()
ZAPI_CLIENT_TOKEN = os.getenv("ZAPI_CLIENT_TOKEN", "").strip()
ZAPI_SEND_URL   = f"https://api.z-api.io/instances/{ZAPI_INSTANCE}/token/{ZAPI_TOKEN}/send-text"
ZAPI_STATUS_URL = f"https://api.z-api.io/instances/{ZAPI_INSTANCE}/token/{ZAPI_TOKEN}/status"
HEADERS = {
    "Client-Token": ZAPI_CLIENT_TOKEN,
    "Content-Type": "application/json",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
}


def so_numeros(t):
    return re.sub(r"\D", "", t or "")


def classificar(tel_bruto):
    """Devolve (telefone_normalizado, motivo_para_pular).

    Celular BR = 55 + DDD(2) + 9 + 8 digitos = 13 chars.

    Numero de 10 digitos (DDD + 8) pode ser duas coisas bem diferentes, e a
    primeira versao deste script tratava as duas como celular antigo, inserindo
    um 9 na frente. Em FIXO isso gera um celular de OUTRA pessoa: em 18/07/2026
    o (75) 2102-3500 virou 5575921023500 e a mensagem foi para um estranho.

    A regra do numero brasileiro separa os dois casos pelo 1o digito:
      - 2 a 5 -> fixo. NAO tem WhatsApp e nao vira celular: pular.
      - 8 ou 9 -> celular antigo, de antes do 9o digito: inserir o 9 e enviar.
    """
    t = so_numeros(tel_bruto)
    if t.startswith("55") and len(t) in (12, 13):
        t = t[2:]  # remove DDI para tratar so a parte nacional

    if len(t) == 10:
        if t[2] in "2345":
            return "", f"fixo ({tel_bruto}) — inserir o 9 mandaria para outra pessoa"
        if t[2] in "89":
            t = t[:2] + "9" + t[2:]   # celular antigo: falta mesmo o 9o digito
        else:
            return "", f"numero fora do padrao ({tel_bruto})"

    if len(t) != 11:
        return "", f"formato inesperado ({len(t)} digitos: {tel_bruto})"
    if t[2] != "9":
        return "", f"nao parece celular ({tel_bruto})"
    return "55" + t, ""


def _req(metodo, url, **kw):
    """requests com fallback: se a verificacao SSL falhar no ambiente local
    (proxy/antivirus interceptando), refaz a chamada sem verificar o certificado."""
    try:
        return requests.request(metodo, url, **kw)
    except requests.exceptions.SSLError:
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        return requests.request(metodo, url, verify=False, **kw)


def zapi_conectada():
    if not (ZAPI_INSTANCE and ZAPI_TOKEN and ZAPI_CLIENT_TOKEN):
        return False, "Z-API nao configurada no .env"
    try:
        r = _req("GET", ZAPI_STATUS_URL, headers={"Client-Token": ZAPI_CLIENT_TOKEN}, timeout=15)
        if r.status_code != 200:
            return False, f"HTTP {r.status_code}: {r.text[:200]}"
        info = r.json()
        if info.get("connected") is True:
            return True, "conectada"
        return False, f"desconectada: {info}"
    except Exception as e:
        return False, str(e)


def enviar(telefone, mensagem):
    try:
        r = _req("POST", ZAPI_SEND_URL, headers=HEADERS,
                 json={"phone": telefone, "message": mensagem}, timeout=30)
        if r.status_code in (200, 201):
            return True, r.text[:200]
        return False, f"HTTP {r.status_code}: {r.text[:200]}"
    except Exception as e:
        return False, str(e)


def ja_enviados_ok():
    """Le o log e retorna set de telefones que ja receberam OK (para nao repetir)."""
    feitos = set()
    if os.path.exists(LOG_CSV):
        with open(LOG_CSV, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                if row.get("status") == "OK":
                    feitos.add(row.get("telefone"))
    return feitos


def carregar_clientes():
    wb = load_workbook(PLANILHA, read_only=True)
    ws = wb.active
    linhas = list(ws.iter_rows(min_row=2, values_only=True))
    clientes = []
    for row in linhas:
        # colunas: ID, Nome, Email, Telefone, Codigo, Data, Marcador
        nome = (row[1] or "").strip()
        tel  = (row[3] or "").strip()
        if not tel:
            continue
        tel_norm, motivo = classificar(tel)
        clientes.append({"nome": nome, "tel_orig": tel,
                         "tel": tel_norm, "pular": motivo})
    return clientes


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--teste", action="store_true", help="Simula, nao envia")
    ap.add_argument("--limite", type=int, default=0, help="Envia so aos N primeiros")
    ap.add_argument("--intervalo", type=int, default=120, help="Segundos entre envios (padrao 120)")
    args = ap.parse_args()

    with open(MSG_FILE, encoding="utf-8") as f:
        template = f.read().strip()

    todos    = carregar_clientes()
    clientes = [c for c in todos if not c["pular"]]
    fora     = [c for c in todos if c["pular"]]
    if args.limite:
        clientes = clientes[: args.limite]

    feitos = ja_enviados_ok()

    print("=" * 60)
    print(f"Planilha : {PLANILHA}")
    print(f"Clientes : {len(todos)}  |  enviaveis: {len(clientes)}  |  fora: {len(fora)}")
    for c in fora:
        print(f"   FORA: {c['nome'] or '(sem nome)'} — {c['pular']}")
    print(f"Ja OK    : {len(feitos)} (serao pulados)")
    print(f"Intervalo: {args.intervalo}s ({args.intervalo/60:.1f} min) entre envios")
    print(f"Modo     : {'TESTE (nao envia)' if args.teste else 'ENVIO REAL'}")
    print("=" * 60)

    if not args.teste:
        ok, msg = zapi_conectada()
        if not ok:
            print(f"[ABORTADO] WhatsApp/Z-API nao conectada: {msg}")
            print("Reconecte em https://app.z-api.io e rode de novo.")
            sys.exit(1)
        print(f"[Z-API] {msg}")

    novo_log = not os.path.exists(LOG_CSV)
    logf = open(LOG_CSV, "a", newline="", encoding="utf-8")
    logw = csv.writer(logf)
    if novo_log:
        logw.writerow(["datahora", "nome", "telefone", "status", "detalhe"])

    total = len(clientes)
    enviados = pulados = falhas = 0

    for i, c in enumerate(clientes, 1):
        nome_completo = c["nome"]
        primeiro_nome = nome_completo.split()[0] if nome_completo else "cliente"
        tel = c["tel"]
        mensagem = template.replace("{nome}", primeiro_nome)
        prefixo = f"[{i}/{total}] {nome_completo or '(sem nome)'} -> {tel}"

        if tel in feitos:
            print(f"{prefixo}  ::  PULADO (ja enviado)")
            pulados += 1
            continue

        agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if args.teste:
            print(f"{prefixo}  ::  [TESTE] enviaria mensagem ({len(mensagem)} chars)")
            logw.writerow([agora, nome_completo, tel, "TESTE", "simulado"])
            logf.flush()
            enviados += 1
        else:
            ok, detalhe = enviar(tel, mensagem)
            status = "OK" if ok else "FALHA"
            print(f"{prefixo}  ::  {status}  {('' if ok else detalhe)}")
            logw.writerow([agora, nome_completo, tel, status, detalhe])
            logf.flush()
            feitos.add(tel) if ok else None
            enviados += 1 if ok else 0
            falhas += 0 if ok else 1

        # espera entre envios (nao espera depois do ultimo)
        if i < total:
            espera = 1 if args.teste else args.intervalo
            if not args.teste:
                print(f"      aguardando {espera}s ({espera/60:.1f} min)...")
            time.sleep(espera)

    logf.close()
    print("=" * 60)
    print(f"Enviados OK: {enviados} | Falhas: {falhas} | Pulados: {pulados}")
    print(f"Log: {LOG_CSV}")


if __name__ == "__main__":
    main()
