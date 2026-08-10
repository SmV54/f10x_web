"""
TESTE — SO LEITURA. Puxa os funcionarios do eSocial (webservices de consulta
de identificadores + download cirurgico) e joga numa planilha .xlsx.

NAO envia nada ao eSocial e NAO grava nada na nossa base:
  - eSocial: so ConsultarIdentificadoresEventos* e SolicitarDownloadEventos*
  - Supabase: so GET (tab_empresa para pegar o certificado, tab_cad para a
    lista de CPFs semente)

Uso:  python _puxar_esocial_funcionarios.py [CNPJ] [tpAmb] [CPF ...]
      CNPJ  default 08777252000133 (FOLHA10 - COMSIST)
      tpAmb 1=Producao (default) / 2=Producao Restrita
      CPF   opcional: so esses CPFs (senao, os da nossa base)

LIMITES DO SERVICO (medidos em 10/08/2026, em producao):
  * 10 solicitacoes por DIA, somando consulta + download (cdResposta 405).
    Estourou, so no dia seguinte. Por isso o script tem orcamento proprio
    (MAX_REQ) e para sozinho quando chega no limite.
  * Consulta de trabalhador: dtIni/dtFim sao xs:dateTime e o intervalo nao
    pode passar de 31 dias (cdResposta 410) - a varredura vai mes a mes.
  * Consulta de empregador so aceita evento sem CPF (S-1260, S-1280, S-1298,
    S-1299, S-5011, S-5012); os outros dao cdResposta 402.
  * Entre os dias 1 e 7 do mes o servico fica bloqueado (cdResposta 403).
"""
import os, re, sys, base64, hashlib, datetime, tempfile, calendar
import requests, urllib3
from dotenv import load_dotenv
from lxml import etree

urllib3.disable_warnings()
load_dotenv(r"C:\folha10-simples\.env")

CNPJ_EMP = re.sub(r"\D", "", sys.argv[1] if len(sys.argv) > 1 else "08777252000133")
TP_AMB   = sys.argv[2] if len(sys.argv) > 2 else "1"

SUPA_URL = os.getenv("SUPABASE_URL").rstrip("/")
SUPA_KEY = os.getenv("SUPABASE_SERVICE_KEY") or os.getenv("SUPABASE_KEY")
H_SUPA   = {"apikey": SUPA_KEY, "Authorization": f"Bearer {SUPA_KEY}"}

OUT_DIR = r"C:\folha10-simples\__teste_esocial"
XML_DIR = os.path.join(OUT_DIR, "xml_brutos")

# ── endpoints (copiados do app.py) ────────────────────────────────────────
BASE_DWN   = "https://webservices.download.esocial.gov.br"
BASE_HOMOL = "https://webservices.producaorestrita.esocial.gov.br"
P_CONS_IDE = "/servicos/empregador/dwlcirurgico/WsConsultarIdentificadoresEventos.svc"
P_DOWNLOAD = "/servicos/empregador/dwlcirurgico/WsSolicitarDownloadEventos.svc"

NS_SOAP = "http://schemas.xmlsoap.org/soap/envelope/"

# As TRES operacoes de consulta ficam no MESMO targetNamespace - conferido no
# WSDL do servico. Usar o namespace "/trabalhador/" ou "/empregador/" aqui da
# SOAP Fault a:ActionNotSupported.
NS_CIE = "http://www.esocial.gov.br/servicos/empregador/consulta/identificadores-eventos/v1_0_0"
NS_DWN = "http://www.esocial.gov.br/servicos/empregador/download/solicitacao/v1_0_0"

_SA_CIE     = NS_CIE + "/ServicoConsultarIdentificadoresEventos/ConsultarIdentificadoresEventos"
SA_CIE_TRAB = _SA_CIE + "Trabalhador"
SA_CIE_EMP  = _SA_CIE + "Empregador"
SA_CIE_TAB  = _SA_CIE + "Tabela"
SA_DWN      = NS_DWN  + "/ServicoSolicitarDownloadEventos/SolicitarDownloadEventosPorId"

# Namespace do XML de DENTRO da solicitacao de download. O "/id/" no fim e
# obrigatorio: sem ele o eSocial responde 417 "targetNamespace invalido".
NS_SCH_DWN_ID = "http://www.esocial.gov.br/schema/download/solicitacao/id/v1_0_0"


def soap_ide_trab(raiz, cpf, dt_ini, dt_fim, tp_evt=""):
    tp = f"<tpEvt>{tp_evt}</tpEvt>" if tp_evt else ""
    body = (f'<eSocial xmlns="http://www.esocial.gov.br/schema/consulta/identificadores-eventos/trabalhador/v1_0_0">'
            f"<consultaIdentificadoresEvts>"
            f"<ideEmpregador><tpInsc>1</tpInsc><nrInsc>{raiz}</nrInsc></ideEmpregador>"
            f"<consultaEvtsTrabalhador><cpfTrab>{cpf}</cpfTrab>"
            f"<dtIni>{dt_ini}</dtIni><dtFim>{dt_fim}</dtFim>{tp}"
            f"</consultaEvtsTrabalhador></consultaIdentificadoresEvts></eSocial>")
    body = assinar(body, PFX, SENHA)
    return (f'<soapenv:Envelope xmlns:soapenv="{NS_SOAP}" xmlns:v1="{NS_CIE}">'
            f"<soapenv:Header/><soapenv:Body>"
            f"<v1:ConsultarIdentificadoresEventosTrabalhador><v1:consultaEventosTrabalhador>{body}"
            f"</v1:consultaEventosTrabalhador></v1:ConsultarIdentificadoresEventosTrabalhador>"
            f"</soapenv:Body></soapenv:Envelope>")


def soap_ide_trab_periodo(raiz, dt_ini, dt_fim, tp_evt=""):
    """SONDAGEM: a MESMA consulta de trabalhador, mas SEM a tag <cpfTrab>.

    Motivo: num cliente novo nao se sabe quem sao os funcionarios, entao
    "todos os eventos de um CPF" nao serve - o que serviria e "todos os eventos
    do mes". Se o schema aceitar a consulta sem cpfTrab, esse e o caminho
    (1 requisicao por mes, cabe folgado na cota de 10/dia).

    RESPONDIDO em 10/08/2026, na Producao Restrita: NAO ACEITA.
        sem cpfTrab -> cd 417 "Erro na estrutura da solicitacao.
                       ...has invalid child element 'dtIni'"
        com cpfTrab -> cd 406 (controle: mesmo envelope, passou)
    Ou seja: nao existe "liste os eventos do mes de todo mundo". A lista de CPFs
    tem que vir de fora - na pratica, do ZIP do portal.
    Fica aqui so como registro; desligado por padrao (ESOCIAL_PROBE_PERIODO=1
    para rodar de novo).
    """
    tp = f"<tpEvt>{tp_evt}</tpEvt>" if tp_evt else ""
    body = (f'<eSocial xmlns="http://www.esocial.gov.br/schema/consulta/identificadores-eventos/trabalhador/v1_0_0">'
            f"<consultaIdentificadoresEvts>"
            f"<ideEmpregador><tpInsc>1</tpInsc><nrInsc>{raiz}</nrInsc></ideEmpregador>"
            f"<consultaEvtsTrabalhador>"
            f"<dtIni>{dt_ini}</dtIni><dtFim>{dt_fim}</dtFim>{tp}"
            f"</consultaEvtsTrabalhador></consultaIdentificadoresEvts></eSocial>")
    body = assinar(body, PFX, SENHA)
    return (f'<soapenv:Envelope xmlns:soapenv="{NS_SOAP}" xmlns:v1="{NS_CIE}">'
            f"<soapenv:Header/><soapenv:Body>"
            f"<v1:ConsultarIdentificadoresEventosTrabalhador><v1:consultaEventosTrabalhador>{body}"
            f"</v1:consultaEventosTrabalhador></v1:ConsultarIdentificadoresEventosTrabalhador>"
            f"</soapenv:Body></soapenv:Envelope>")


def soap_ide_emp(raiz, tp_evt, per_apur):
    body = (f'<eSocial xmlns="http://www.esocial.gov.br/schema/consulta/identificadores-eventos/empregador/v1_0_0">'
            f"<consultaIdentificadoresEvts>"
            f"<ideEmpregador><tpInsc>1</tpInsc><nrInsc>{raiz}</nrInsc></ideEmpregador>"
            f"<consultaEvtsEmpregador><tpEvt>{tp_evt}</tpEvt><perApur>{per_apur}</perApur>"
            f"</consultaEvtsEmpregador></consultaIdentificadoresEvts></eSocial>")
    body = assinar(body, PFX, SENHA)
    return (f'<soapenv:Envelope xmlns:soapenv="{NS_SOAP}" xmlns:v1="{NS_CIE}">'
            f"<soapenv:Header/><soapenv:Body>"
            f"<v1:ConsultarIdentificadoresEventosEmpregador><v1:consultaEventosEmpregador>{body}"
            f"</v1:consultaEventosEmpregador></v1:ConsultarIdentificadoresEventosEmpregador>"
            f"</soapenv:Body></soapenv:Envelope>")


def soap_ide_tab(raiz, tp_evt):
    body = (f'<eSocial xmlns="http://www.esocial.gov.br/schema/consulta/identificadores-eventos/tabela/v1_0_0">'
            f"<consultaIdentificadoresEvts>"
            f"<ideEmpregador><tpInsc>1</tpInsc><nrInsc>{raiz}</nrInsc></ideEmpregador>"
            f"<consultaEvtsTabela><tpEvt>{tp_evt}</tpEvt></consultaEvtsTabela>"
            f"</consultaIdentificadoresEvts></eSocial>")
    body = assinar(body, PFX, SENHA)
    return (f'<soapenv:Envelope xmlns:soapenv="{NS_SOAP}" xmlns:v1="{NS_CIE}">'
            f"<soapenv:Header/><soapenv:Body>"
            f"<v1:ConsultarIdentificadoresEventosTabela><v1:consultaEventosTabela>{body}"
            f"</v1:consultaEventosTabela></v1:ConsultarIdentificadoresEventosTabela>"
            f"</soapenv:Body></soapenv:Envelope>")


def soap_download(raiz, ids):
    ids_xml = "".join(f"<id>{i}</id>" for i in ids[:40])
    body = (f'<eSocial xmlns="{NS_SCH_DWN_ID}">'
            f"<download><ideEmpregador><tpInsc>1</tpInsc><nrInsc>{raiz}</nrInsc></ideEmpregador>"
            f"<solicDownloadEvtsPorId>{ids_xml}</solicDownloadEvtsPorId></download></eSocial>")
    body = assinar(body, PFX, SENHA)
    return (f'<soapenv:Envelope xmlns:soapenv="{NS_SOAP}" xmlns:ser="{NS_DWN}">'
            f"<soapenv:Header/><soapenv:Body>"
            f"<ser:SolicitarDownloadEventosPorId><ser:solicitacao>{body}"
            f"</ser:solicitacao></ser:SolicitarDownloadEventosPorId>"
            f"</soapenv:Body></soapenv:Envelope>")


def assinar(xml_str, pfx_bytes, senha):
    """XMLDSig enveloped com URI="" — o eSocial EXIGE a consulta assinada
    (sem isso responde 417 'incomplete content ... expected Signature')."""
    from signxml import XMLSigner, methods
    from copy import deepcopy
    from cryptography.hazmat.primitives import hashes as _h
    from cryptography.hazmat.primitives.asymmetric import padding as _p
    from cryptography.hazmat.primitives.serialization import (
        pkcs12, Encoding, PrivateFormat, NoEncryption)
    pk, cert, _ = pkcs12.load_key_and_certificates(
        pfx_bytes, senha.encode() if isinstance(senha, str) else senha)
    key_pem  = pk.private_bytes(Encoding.PEM, PrivateFormat.TraditionalOpenSSL, NoEncryption())
    cert_pem = cert.public_bytes(Encoding.PEM)
    root = etree.fromstring(xml_str.encode("utf-8"))
    filho = root[0]
    tirar_id = not filho.get("Id")
    if tirar_id:
        filho.set("Id", "ID1")
    signer = XMLSigner(method=methods.enveloped, signature_algorithm="rsa-sha256",
                       digest_algorithm="sha256",
                       c14n_algorithm="http://www.w3.org/TR/2001/REC-xml-c14n-20010315")
    signed = signer.sign(root, key=key_pem, cert=cert_pem, reference_uri=f"#{filho.get('Id')}")
    NS_DS = "http://www.w3.org/2000/09/xmldsig#"
    ref = signed.find(f".//{{{NS_DS}}}Reference")
    ref.set("URI", "")
    if tirar_id:
        for ch in signed:
            if lname(ch) != "Signature" and ch.get("Id") == "ID1":
                del ch.attrib["Id"]
    sem = deepcopy(signed)
    s = sem.find(f"{{{NS_DS}}}Signature")
    if s is not None:
        sem.remove(s)
    ref.find(f"{{{NS_DS}}}DigestValue").text = base64.b64encode(
        hashlib.sha256(etree.tostring(sem, method="c14n")).digest()).decode()
    si = signed.find(f".//{{{NS_DS}}}SignedInfo")
    signed.find(f".//{{{NS_DS}}}SignatureValue").text = base64.b64encode(
        pk.sign(etree.tostring(si, method="c14n"), _p.PKCS1v15(), _h.SHA256())).decode()
    return etree.tostring(signed, encoding="unicode", xml_declaration=False)


# O servico so aceita 10 solicitacoes por dia (consulta + download juntos).
# Estas variaveis seguram o script para nao torrar a cota sem querer.
MAX_REQ  = int(os.getenv("ESOCIAL_MAX_REQ", "9"))
REQ_FEITAS = 0
COTA_ESTOUROU = False


class SemCota(Exception):
    pass


def gastar(quantas=1):
    """Levanta SemCota quando o orcamento do dia acabou."""
    global REQ_FEITAS
    if COTA_ESTOUROU:
        raise SemCota("o eSocial ja respondeu 405 (10 solicitacoes/dia) - so amanha")
    if REQ_FEITAS + quantas > MAX_REQ:
        raise SemCota(f"orcamento do script esgotado ({REQ_FEITAS}/{MAX_REQ} requisicoes)")
    REQ_FEITAS += quantas


def marcar_405(cd, desc):
    """Marca a cota como estourada quando o eSocial devolve 405."""
    global COTA_ESTOUROU
    if cd == "405" or "solicitacoes por dia" in (desc or "") or "solicitações por dia" in (desc or ""):
        COTA_ESTOUROU = True
        return True
    return False


def http_post_cert(url, soap, pfx_bytes, senha, soap_action=""):
    from cryptography.hazmat.primitives.serialization import (
        pkcs12, Encoding, PrivateFormat, NoEncryption)
    pk, cert, _ = pkcs12.load_key_and_certificates(
        pfx_bytes, senha.encode() if isinstance(senha, str) else senha)
    kf = tempfile.NamedTemporaryFile(delete=False, suffix=".key.pem")
    cf = tempfile.NamedTemporaryFile(delete=False, suffix=".cert.pem")
    try:
        kf.write(pk.private_bytes(Encoding.PEM, PrivateFormat.TraditionalOpenSSL, NoEncryption())); kf.close()
        cf.write(cert.public_bytes(Encoding.PEM)); cf.close()
        resp = requests.post(url, data=soap.encode("utf-8"),
                             headers={"Content-Type": "text/xml;charset=UTF-8",
                                      "SOAPAction": f'"{soap_action}"'},
                             cert=(cf.name, kf.name), verify=False, timeout=90)
        if resp.status_code != 200 or not (resp.text or "").strip():
            raise Exception(f"HTTP {resp.status_code}: {(resp.text or '')[:400]}")
        return resp.text
    finally:
        os.unlink(kf.name); os.unlink(cf.name)


def lname(el):
    return el.tag.split("}")[-1] if isinstance(el.tag, str) else ""


def parse_ide(xml_resp):
    out = {"ok": False, "cd": "", "desc": "", "ids": [], "erro": ""}
    try:
        root = etree.fromstring(xml_resp.encode("utf-8", errors="replace"))
    except Exception as e:
        out["erro"] = f"XML invalido: {e}"; return out
    for el in root.iter():
        if lname(el) == "faultstring":
            out["erro"] = f"SOAP Fault: {(el.text or '').strip()}"; return out
    for el in root.iter():
        t, txt = lname(el), (el.text or "").strip()
        if t == "cdResposta":   out["cd"] = txt
        elif t == "descResposta": out["desc"] = txt
    # A resposta traz <identificadoresEvts><identificadorEvt><id/><nrRec/>.
    # O nome "ideEvento" (usado antes) nao existe - por isso a lista vinha vazia
    # mesmo com o cdResposta dizendo que achou eventos.
    for el in root.iter():
        if lname(el) != "identificadorEvt":
            continue
        item = {"id": "", "tpEvt": "", "nrRecEvt": "", "perApur": ""}
        for c in el:
            t, txt = lname(c), (c.text or "").strip()
            if t == "nrRec":
                item["nrRecEvt"] = txt
            elif t in item:
                item[t] = txt
        if item["id"]:
            out["ids"].append(item)
    out["ok"] = out["cd"].startswith("2") if out["cd"] else bool(out["ids"])
    return out


def parse_dwn(xml_resp):
    out = {"ok": False, "cd": "", "desc": "", "eventos": [], "erro": ""}
    try:
        root = etree.fromstring(xml_resp.encode("utf-8", errors="replace"))
    except Exception as e:
        out["erro"] = f"XML invalido: {e}"; return out
    for el in root.iter():
        if lname(el) == "faultstring":
            out["erro"] = f"SOAP Fault: {(el.text or '').strip()}"; return out
    for el in root.iter():
        t, txt = lname(el), (el.text or "").strip()
        if t == "cdResposta":   out["cd"] = txt
        elif t == "descResposta": out["desc"] = txt
    for el in root.iter():
        if lname(el) != "evento":
            continue
        evt_id, inner = el.get("Id") or el.get("id") or "", ""
        for c in el:
            if isinstance(c.tag, str):
                inner = etree.tostring(c, encoding="unicode"); break
        if inner:
            out["eventos"].append({"id": evt_id, "xml": inner})
    out["ok"] = out["cd"].startswith("2") if out["cd"] else bool(out["eventos"])
    return out


# ── 1. Empresa + certificado (leitura no Supabase) ────────────────────────
print("=" * 72)
print(f"TESTE eSocial — CNPJ {CNPJ_EMP} — ambiente {'Producao' if TP_AMB=='1' else 'Producao Restrita'}")
print("=" * 72)

r = requests.get(f"{SUPA_URL}/rest/v1/tab_empresa",
                 params={"select": "*", "cnpj": f"eq.{CNPJ_EMP}"},
                 headers=H_SUPA, verify=False, timeout=30)
emp = (r.json() or [None])[0]
if not emp:
    print("!! Empresa nao encontrada em tab_empresa"); sys.exit(1)
print(f"Empresa: {emp.get('razaosocial') or emp.get('nome_fantasia')} (id_empresa={emp.get('id_empresa')})")

def cert_decrypt(token):
    from cryptography.fernet import Fernet
    raw = (os.getenv("FLASK_SECRET_KEY", "F10default") + "_cert_v1").encode()
    return Fernet(base64.urlsafe_b64encode(hashlib.sha256(raw).digest())).decrypt(token.encode()).decode()

try:
    PFX   = base64.b64decode(emp["cert_pfx_b64"])
    SENHA = cert_decrypt(emp["cert_senha_enc"])
except Exception as e:
    print(f"!! Nao consegui abrir o certificado: {e}")
    print("   (a senha so descriptografa no ambiente onde o .pfx foi gravado)")
    sys.exit(1)

from cryptography.hazmat.primitives.serialization import pkcs12
_pk, _c, _ = pkcs12.load_key_and_certificates(PFX, SENHA.encode())
print(f"Certificado OK — titular: {_c.subject.rfc4514_string()[:70]}")
print(f"                 validade ate: {_c.not_valid_after_utc.strftime('%d/%m/%Y')}")

RAIZ = CNPJ_EMP[:8]
BASE = BASE_DWN if TP_AMB == "1" else BASE_HOMOL
URL_IDE, URL_DWN = BASE + P_CONS_IDE, BASE + P_DOWNLOAD
os.makedirs(XML_DIR, exist_ok=True)

diag = []   # linhas da aba Diagnostico
print(f"Orcamento deste run: {MAX_REQ} requisicao(oes)  "
      f"(o eSocial so aceita 10 por dia, consulta + download juntos)")

# ── 2. TESTE A — conectividade (evento de tabela S-1000, nao precisa CPF) ──
print("\n[A] Consulta de identificadores — S-1000 (tabela)")
BLOQUEADO = False
FORA_DO_AR = False
try:
    gastar()
    p = parse_ide(http_post_cert(URL_IDE, soap_ide_tab(RAIZ, "S-1000"), PFX, SENHA, SA_CIE_TAB))
    print(f"    cd={p['cd']} {p['desc'][:110]} | eventos: {len(p['ids'])} | erro: {p['erro'][:120]}")
    diag.append(["A", "S-1000 (tabela)", "-", p["cd"], p["desc"], len(p["ids"]), p["erro"]])
    if marcar_405(p["cd"], p["desc"]):
        print("    >>> A cota de 10 solicitacoes/dia acabou. Rode de novo amanha.")
    if p["cd"] == "403" or "entre os dias 1 e 7" in (p["desc"] or ""):
        BLOQUEADO = True
except SemCota as e:
    print(f"    (pulado: {e})")
except Exception as e:
    print(f"    !! {str(e)[:300]}")
    diag.append(["A", "S-1000 (tabela)", "-", "", "", 0, str(e)[:300]])
    if "HTTP 500" in str(e):
        # Dois 500 DIFERENTES chegam aqui - nao confundir:
        #  * corpo VAZIO  -> e a janela dos dias 1 a 7 (a Producao Restrita, com a
        #    MESMA requisicao, responde cd=403 com a mensagem legivel).
        #  * corpo HTML   -> o servico dwlcirurgico caiu do lado do governo
        #    ("Server Error in '/servicos/empregador/dwlcirurgico' Application").
        #    Visto em 10/08/2026: Unable to open configSource file
        #    'Configs\\environmentSettings.config'. Nao adianta tentar de novo hoje
        #    e NAO tem nada a ver com a janela do mes.
        if "<html" in str(e).lower() or "Server Error" in str(e):
            FORA_DO_AR = True
            diag.append(["A", "diagnostico", "-", "500",
                         "Servico dwlcirurgico fora do ar (erro de configuracao do governo)",
                         0, str(e)[:300]])
        else:
            BLOQUEADO = True
            diag.append(["A", "diagnostico", "-", "500",
                         "Producao devolveu 500 vazio; restrita responde cd=403 (janela dia 1-7)",
                         0, ""])

if FORA_DO_AR:
    print("\n    >>> O servico dwlcirurgico esta FORA DO AR (erro do proprio eSocial).")
    print("    >>> Nao e cota, nao e certificado, nao e a janela do mes. Tentar mais tarde.")
elif BLOQUEADO:
    print("\n    >>> O eSocial bloqueia solicitacoes de download entre os dias 1 e 7 do mes.")
    print("    >>> Nada a puxar hoje; rode de novo a partir do dia 8.")
BLOQUEADO = BLOQUEADO or FORA_DO_AR

# ── 2b. TESTE A2 — consulta de trabalhador SEM cpfTrab (so o periodo) ─────
# A pergunta que decide a tela do cliente novo: da para pedir "todos os eventos
# do mes" sem saber os CPFs? Custa 1 requisicao e responde de uma vez.
hoje = datetime.date.today()
_mes_sonda = os.getenv("ESOCIAL_MES_SONDA") or (
    hoje.replace(day=1) - datetime.timedelta(days=1)).strftime("%Y-%m")
IDS_PERIODO = []
if os.getenv("ESOCIAL_PROBE_PERIODO", "0") == "1" and not BLOQUEADO:
    _a, _m = int(_mes_sonda[:4]), int(_mes_sonda[5:7])
    _ult = calendar.monthrange(_a, _m)[1]
    print(f"\n[A2] Consulta de TRABALHADOR sem cpfTrab — mes {_mes_sonda}")
    try:
        gastar()
        p = parse_ide(http_post_cert(
            URL_IDE,
            soap_ide_trab_periodo(RAIZ, f"{_a:04d}-{_m:02d}-01T00:00:00",
                                  f"{_a:04d}-{_m:02d}-{_ult:02d}T23:59:59"),
            PFX, SENHA, SA_CIE_TRAB))
        print(f"    cd={p['cd']} {p['desc'][:110]} | eventos: {len(p['ids'])} | erro: {p['erro'][:150]}")
        diag.append(["A2", "trabalhador sem cpfTrab", _mes_sonda, p["cd"], p["desc"],
                     len(p["ids"]), p["erro"][:300]])
        marcar_405(p["cd"], p["desc"])
        if p["ids"]:
            IDS_PERIODO = p["ids"]
            print("    >>> ACEITOU sem CPF. E este o caminho para o cliente novo:")
            print("    >>> 1 consulta por mes devolve os eventos de todos os trabalhadores.")
        elif p["erro"] or (p["cd"] and not p["cd"].startswith("2")):
            print("    >>> Recusou. cpfTrab e obrigatorio mesmo; a lista de CPFs tem que")
            print("    >>> vir por fora (importacao de XML / portal do eSocial).")
    except SemCota as e:
        print(f"    (pulado: {e})")
    except Exception as e:
        print(f"    !! {str(e)[:300]}")
        diag.append(["A2", "trabalhador sem cpfTrab", _mes_sonda, "", "", 0, str(e)[:300]])

# ── 3. TESTE B — consulta do EMPREGADOR (eventos sem CPF) ─────────────────
# So estes tpEvt sao aceitos aqui; qualquer outro devolve cd=402
# "Solicitacao invalida" (testado um a um em 10/08/2026).
TP_EMPREGADOR = ("S-1299", "S-1260", "S-1280", "S-1298", "S-5011", "S-5012")
per_ini =(hoje.replace(day=1) - datetime.timedelta(days=1)).strftime("%Y-%m")
print(f"\n[B] Consulta do EMPREGADOR (sem CPF) — perApur {per_ini}")
ids_empregador = []
for tp in (() if BLOQUEADO else TP_EMPREGADOR):
    try:
        gastar()
        p = parse_ide(http_post_cert(URL_IDE, soap_ide_emp(RAIZ, tp, per_ini), PFX, SENHA, SA_CIE_EMP))
        print(f"    {tp}: cd={p['cd']} | eventos: {len(p['ids'])} | {p['desc'][:70]}{p['erro'][:90]}")
        diag.append(["B", tp, per_ini, p["cd"], p["desc"], len(p["ids"]), p["erro"]])
        if marcar_405(p["cd"], p["desc"]):
            print("    >>> cota do dia esgotada."); break
        for i in p["ids"]:
            i["tpEvtCons"] = tp
        ids_empregador.extend(p["ids"])
    except SemCota as e:
        print(f"    {tp}: (pulado: {e})"); break
    except Exception as e:
        print(f"    {tp}: !! {str(e)[:200]}")
        diag.append(["B", tp, per_ini, "", "", 0, str(e)[:300]])
diag.append(["B", "TOTAL ids do empregador", per_ini, "", "", len(ids_empregador), ""])

# ── 4. Lista de CPFs ──────────────────────────────────────────────────────
# cpfTrab e obrigatorio no schema: nao existe "liste todos os empregados".
# Ou vem por parametro, ou sai da nossa base.
CPFS = [re.sub(r"\D", "", c) for c in sys.argv[3:] if len(re.sub(r"\D", "", c)) == 11]
if not CPFS:
    r = requests.get(f"{SUPA_URL}/rest/v1/tab_cad",
                     params={"select": "cpf,nome,matricula,situacao",
                             "id_empresa": f"eq.{emp['id_empresa']}", "limit": "500"},
                     headers=H_SUPA, verify=False, timeout=30)
    base_rows = r.json() if r.status_code == 200 else []
    CPFS = sorted({re.sub(r"\D", "", c.get("cpf") or "") for c in base_rows} - {""})
    CPFS = [c for c in CPFS if len(c) == 11]
print(f"\n[C] {len(CPFS)} CPF(s) a consultar")

# ── 5. Consulta dos eventos de cada trabalhador ───────────────────────────
# dtIni/dtFim sao xs:dateTime E o intervalo nao pode passar de 31 dias
# (cd=410), entao a varredura vai de mes em mes. Sem tpEvt vem tudo o que
# houver do trabalhador na janela.
ANO_INI = int(os.getenv("ESOCIAL_ANO_INI", str(hoje.year - 1)))


def janelas_mensais(ano_ini, ate):
    ano, mes = ano_ini, 1
    while (ano, mes) <= (ate.year, ate.month):
        ult = calendar.monthrange(ano, mes)[1]
        yield (f"{ano:04d}-{mes:02d}-01T00:00:00", f"{ano:04d}-{mes:02d}-{ult:02d}T23:59:59",
               f"{ano:04d}-{mes:02d}")
        mes += 1
        if mes > 12:
            ano, mes = ano + 1, 1


eventos_todos, xml_por_id = [], {}
parou_em = None
for cpf in ([] if BLOQUEADO else CPFS):
    achados = []
    for dt_ini, dt_fim, rot in janelas_mensais(ANO_INI, hoje):
        try:
            gastar()
        except SemCota as e:
            parou_em = (cpf, rot, str(e))
            break
        try:
            p = parse_ide(http_post_cert(URL_IDE, soap_ide_trab(RAIZ, cpf, dt_ini, dt_fim),
                                         PFX, SENHA, SA_CIE_TRAB))
        except Exception as e:
            diag.append(["C", "janela " + rot, cpf, "", "", 0, str(e)[:200]])
            continue
        if marcar_405(p["cd"], p["desc"]):
            parou_em = (cpf, rot, "405 - cota de 10/dia esgotada")
            break
        if p["ids"]:
            diag.append(["C", "janela " + rot, cpf, p["cd"], p["desc"], len(p["ids"]), ""])
            for it in p["ids"]:
                it["cpf"], it["tpEvtCons"] = cpf, rot
            achados.extend(p["ids"])
    eventos_todos.extend(achados)
    print(f"    {cpf}: {len(achados)} evento(s)")
    if parou_em:
        break

# ── 5b. Download dos eventos achados (cada chamada leva ate 40 ids) ───────
TODOS_IDS = [e["id"] for e in eventos_todos] + [i["id"] for i in ids_empregador]
for i in range(0, len(TODOS_IDS), 40):
    lote = TODOS_IDS[i:i + 40]
    try:
        gastar()
    except SemCota as e:
        print(f"    download: (pulado: {e})")
        break
    try:
        d = parse_dwn(http_post_cert(URL_DWN, soap_download(RAIZ, lote), PFX, SENHA, SA_DWN))
        if marcar_405(d["cd"], d["desc"]):
            print("    download: cota do dia esgotada."); break
        if d["erro"]:
            diag.append(["D", "download", f"lote {i//40+1}", "", "", 0, d["erro"][:200]]); continue
        diag.append(["D", "download", f"lote {i//40+1}", d["cd"], d["desc"], len(d["eventos"]), ""])
        for ev in d["eventos"]:
            xml_por_id[ev["id"]] = ev["xml"]
            nome = re.sub(r"[^A-Za-z0-9_.-]", "_", ev["id"])[:60] or f"evt{i}"
            with open(os.path.join(XML_DIR, nome + ".xml"), "w", encoding="utf-8") as f:
                f.write(ev["xml"])
    except Exception as e:
        diag.append(["D", "download", f"lote {i//40+1}", "", "", 0, str(e)[:200]])

print(f"\n    Total de eventos localizados: {len(eventos_todos) + len(ids_empregador)} "
      f"| XMLs baixados: {len(xml_por_id)}")
print(f"    Requisicoes gastas: {REQ_FEITAS}")
if parou_em:
    print(f"    PAROU em cpf {parou_em[0]}, janela {parou_em[1]} — {parou_em[2]}")
    print(f"    Retome amanha a partir dai (ESOCIAL_ANO_INI={parou_em[1][:4]}).")
    diag.append(["C", "PAROU", parou_em[0], "", parou_em[2], 0, "janela " + parou_em[1]])

# ── 6. Monta a planilha ───────────────────────────────────────────────────
def txt(root, tag):
    for el in root.iter():
        if lname(el) == tag and (el.text or "").strip():
            return (el.text or "").strip()
    return ""

func = {}   # cpf -> dict
for ev_id, xml in xml_por_id.items():
    try:
        root = etree.fromstring(xml.encode("utf-8"))
    except Exception:
        continue
    tipo = ""
    for el in root.iter():
        if lname(el).startswith("evt"):
            tipo = lname(el); break
    cpf = txt(root, "cpfTrab") or txt(root, "cpfBenef")
    if not cpf:
        continue
    f = func.setdefault(cpf, {"cpf": cpf})
    if tipo in ("evtAdmissao", "evtTSVInicio", "evtCadInicial"):
        f.update({
            "nome":      txt(root, "nmTrab"),
            "nascto":    txt(root, "dtNascto"),
            "sexo":      txt(root, "sexo"),
            "matricula": txt(root, "matricula"),
            "admissao":  txt(root, "dtAdm") or txt(root, "dtInicio"),
            "categoria": txt(root, "codCateg"),
            "cargo":     txt(root, "nmCargo"),
            "cbo":       txt(root, "CBOCargo") or txt(root, "codCBO"),
            "salario":   txt(root, "vrSalFx"),
            "und_sal":   txt(root, "undSalFixo"),
            "tp_contr":  txt(root, "tpContr"),
            "hrs_sem":   txt(root, "qtdHrsSem"),
            "tp_reg_trab": txt(root, "tpRegTrab"),
            "sind_cnpj": txt(root, "cnpjSindCategProf"),
            "cep":       txt(root, "cep"),
            "municipio": txt(root, "codMunic"),
            "uf":        txt(root, "uf"),
            "evento":    tipo,
        })
    elif tipo in ("evtAltContratual", "evtAltCadastral"):
        for k, tag in (("salario_alt", "vrSalFx"), ("cargo_alt", "nmCargo"),
                       ("cbo_alt", "CBOCargo"), ("dt_alt", "dtAlteracao")):
            v = txt(root, tag)
            if v: f[k] = v
        f.setdefault("nome", txt(root, "nmTrab"))
    elif tipo in ("evtDeslig", "evtTSVTermino"):
        f["desligamento"] = txt(root, "dtDeslig") or txt(root, "dtTerm")
        f["mtv_deslig"]   = txt(root, "mtvDeslig")

COLS = [("cpf", "CPF"), ("nome", "Nome"), ("nascto", "Nascimento"), ("sexo", "Sexo"),
        ("matricula", "Matricula"), ("admissao", "Admissao"), ("categoria", "Categoria"),
        ("cargo", "Cargo"), ("cbo", "CBO"), ("salario", "Salario"), ("und_sal", "Und.Sal"),
        ("tp_contr", "Tp.Contrato"), ("hrs_sem", "Hrs/Sem"), ("tp_reg_trab", "Reg.Trab"),
        ("sind_cnpj", "CNPJ Sindicato"), ("cep", "CEP"), ("municipio", "Cod.Munic"), ("uf", "UF"),
        ("cargo_alt", "Cargo (S-2206)"), ("salario_alt", "Salario (S-2206)"), ("dt_alt", "Dt.Alteracao"),
        ("desligamento", "Desligamento"), ("mtv_deslig", "Mtv.Deslig"), ("evento", "Origem")]

import openpyxl
from openpyxl.styles import Font, PatternFill
wb = openpyxl.Workbook()

ws = wb.active; ws.title = "Funcionarios"
ws.append([c[1] for c in COLS])
for cpf in sorted(func):
    ws.append([func[cpf].get(k, "") for k, _ in COLS])

ws2 = wb.create_sheet("Eventos")
ws2.append(["CPF", "tpEvt consultado", "tpEvt retornado", "ID do evento", "Recibo", "Baixado?"])
for e in eventos_todos:
    ws2.append([e.get("cpf", ""), e.get("tpEvtCons", ""), e.get("tpEvt", ""),
                e.get("id", ""), e.get("nrRecEvt", ""), "SIM" if e.get("id") in xml_por_id else "nao"])

ws3 = wb.create_sheet("Diagnostico")
ws3.append(["Teste", "Evento", "Chave", "cdResposta", "descResposta", "Qtd", "Erro"])
for d in diag:
    ws3.append(d)

for sh in (ws, ws2, ws3):
    for c in sh[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
    sh.freeze_panes = "A2"
    for col in sh.columns:
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        sh.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 45)

os.makedirs(OUT_DIR, exist_ok=True)
stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M")
out = os.path.join(OUT_DIR, f"Funcionarios_eSocial_{CNPJ_EMP}_{stamp}.xlsx")
wb.save(out)
print(f"\nPLANILHA: {out}")
print(f"XMLs brutos (so para conferencia): {XML_DIR}")
print(f"Funcionarios montados: {len(func)}")
