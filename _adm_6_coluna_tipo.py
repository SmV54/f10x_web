# -*- coding: utf-8 -*-
"""Cria a coluna "Tipo de contato" na planilha comercial e marca o que ja existe.

Coluna 9, depois de DataUltimaMensagem/UltimaMensagem, para nao mexer nas 1 a 8
que o enviar_whatsapp_sindicos.py le. Tudo que ja esta na planilha veio do ramo
condominial, entao entra como "Sindico/Administradora".

Uso: python _adm_6_coluna_tipo.py
"""
import os, re, sys, shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

sys.stdout.reconfigure(encoding="utf-8")

BASE     = os.path.dirname(os.path.abspath(__file__))
PLANILHA = os.path.join(BASE, "Comercial", "Sindicos_Profissionais_Brasil.xlsx")
LINHA_CABECALHO = 3
PRIMEIRA_LINHA  = 4
COL_TIPO        = 9
TIPO_PADRAO     = "Sindico/Administradora"

wb = load_workbook(PLANILHA)
ws = wb.worksheets[0]

atual = str(ws.cell(row=LINHA_CABECALHO, column=COL_TIPO).value or "").strip()
if atual and atual != "Tipo de contato":
    print(f"A coluna {COL_TIPO} ja tem outro cabecalho: {atual!r}. Nada feito.")
    sys.exit(1)

bkp = PLANILHA.replace(".xlsx", f"_bkp_{datetime.now():%Y%m%d-%H%M%S}.xlsx")
shutil.copy2(PLANILHA, bkp)
print("Copia de seguranca:", bkp)

modelo = ws.cell(row=LINHA_CABECALHO, column=1)
c = ws.cell(row=LINHA_CABECALHO, column=COL_TIPO, value="Tipo de contato")
if modelo.font:
    c.font = Font(bold=modelo.font.bold, size=modelo.font.size, name=modelo.font.name)

marcados = 0
for r in range(PRIMEIRA_LINHA, ws.max_row + 1):
    if not str(ws.cell(row=r, column=3).value or "").strip():
        continue
    if not str(ws.cell(row=r, column=COL_TIPO).value or "").strip():
        ws.cell(row=r, column=COL_TIPO, value=TIPO_PADRAO)
        marcados += 1

ws.column_dimensions["I"].width = 24
wb.save(PLANILHA)
print(f"Coluna 'Tipo de contato' criada. {marcados} linhas marcadas como {TIPO_PADRAO!r}.")
