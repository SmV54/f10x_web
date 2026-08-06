# -*- coding: utf-8 -*-
"""ETAPA 3 - junta o que foi coletado e inclui na planilha comercial.

Entradas:
  Comercial\\_adm_fichas.json          (etapa 2 - telefone declarado na ficha)
  Comercial\\_coleta_adm_sites.csv     (coletor rodado nos sites das empresas)

Regras herdadas da campanha de sindicos, pelo mesmo motivo: numero errado =
mensagem promocional para um estranho.
  - so celular (11 digitos, 9 na frente, proximo digito 6-9, nao repetitivo)
  - um contato por pagina, ficando com a evidencia mais forte
  - telefone que ja esta na planilha nao entra de novo

Forca da evidencia: wa.me > ficha (telefone declarado no cadastro do
diretorio) > tel: > texto solto na pagina.

Uso:  python _adm_3_planilha.py            (mostra o que entraria)
      python _adm_3_planilha.py --gravar   (grava na planilha, com backup)
"""
import os, re, csv, sys, json, shutil, unicodedata
from datetime import datetime

from openpyxl import load_workbook

BASE     = os.path.dirname(os.path.abspath(__file__))
FICHAS   = os.path.join(BASE, "Comercial", "_adm_fichas.json")
CSV_SITE = os.path.join(BASE, "Comercial", "_coleta_adm_sites.csv")
PLANILHA = os.path.join(BASE, "Comercial", "Sindicos_Profissionais_Brasil.xlsx")

PRIMEIRA_LINHA = 4
FORCA = {"wa.me": 4, "ficha": 3, "tel:": 2, "texto": 1}

DDDS = {
    "11","12","13","14","15","16","17","18","19","21","22","24","27","28",
    "31","32","33","34","35","37","38","41","42","43","44","45","46","47","48","49",
    "51","53","54","55","61","62","63","64","65","66","67","68","69",
    "71","73","74","75","77","79","81","82","83","84","85","86","87","88","89",
    "91","92","93","94","95","96","97","98","99",
}
UF_POR_DDD = {
    "11":"SP","12":"SP","13":"SP","14":"SP","15":"SP","16":"SP","17":"SP","18":"SP","19":"SP",
    "21":"RJ","22":"RJ","24":"RJ","27":"ES","28":"ES",
    "31":"MG","32":"MG","33":"MG","34":"MG","35":"MG","37":"MG","38":"MG",
    "41":"PR","42":"PR","43":"PR","44":"PR","45":"PR","46":"PR",
    "47":"SC","48":"SC","49":"SC","51":"RS","53":"RS","54":"RS","55":"RS",
    "61":"DF","62":"GO","63":"TO","64":"GO","65":"MT","66":"MT","67":"MS","68":"AC","69":"RO",
    "71":"BA","73":"BA","74":"BA","75":"BA","77":"BA","79":"SE",
    "81":"PE","82":"AL","83":"PB","84":"RN","85":"CE","86":"PI","87":"PE","88":"CE","89":"PI",
    "91":"PA","92":"AM","93":"PA","94":"PA","95":"RR","96":"AP","97":"AM","98":"MA","99":"MA",
}

LIXO = re.compile(r"(p[aá]gina n[aã]o encontrada|not found|404|erro|error|"
                  r"dom[ií]nio|domain|em constru[cç][aã]o|hospedagem|registro\.br)", re.I)

# Nome generico demais para identificar a empresa na hora de enviar
GENERICO = re.compile(r"^(administrador|administradora|administra[cç][aã]o|"
                      r"condom[ií]nios?|gest[aã]o|s[ií]ndico profissional|"
                      r"s[ií]ndica profissional|home|in[ií]cio|contato)$", re.I)

# Confirma o ramo pelo titulo do site (usado so nas listas de associadas)
RE_RAMO = re.compile(r"condom[ií]nio|condominial|s[ií]ndic|predial|"
                     r"administra[cç][aã]o de bens|gest[aã]o condominial", re.I)

# Outro ramo que se cadastrou na categoria de administradora (acontece muito
# no diretorio). So derruba quando o nome NAO fala de condominio, senao
# "Consultoria Condominial" e "Contabilidade Condominial" cairiam junto.
RE_OUTRO_RAMO = re.compile(r"gerador|elevador|advocacia|advogad|jur[ií]dic|"
                           r"seguros|dedetiz|desentup|limpeza|uniforme|c[aâ]mera|"
                           r"monitoramento|energia solar|climatiza|jardinagem|"
                           r"portaria remota|software|aplicativo|marketing|"
                           r"constru[tç]|engenharia|per[ií]cia|perito|"
                           r"recursos humanos|treinamento|\bti\b|tecnologia|"
                           r"inform[áa]tica|seguran[çc]a|auditoria empresarial", re.I)


def fora_do_ramo(nome):
    return bool(RE_OUTRO_RAMO.search(nome or "")) and not RE_RAMO.search(nome or "")


def normalizar(bruto):
    """Devolve 55DD9XXXXXXXX, ou "" se nao for celular valido."""
    t = re.sub(r"\D", "", str(bruto or ""))
    if t.startswith("55") and len(t) in (12, 13):
        t = t[2:]
    if len(t) != 11 or t[2] != "9":
        return ""
    if t[:2] not in DDDS or t[3] not in "6789":
        return ""
    if len(set(t[2:])) <= 2:
        return ""
    return "55" + t


def formatar(tel):
    n = tel[2:]
    return f"({n[:2]}) {n[2:7]}-{n[7:]}"


def limpar_nome(t):
    t = re.sub(r"\s+", " ", t or "").strip()
    if "Ã" in t or "Â" in t:
        try:
            t = t.encode("latin-1").decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            pass
    t = re.sub(r"^(Contato|Home|Início|Inicio|Quem somos)\s*[-–—|:]\s*", "", t, flags=re.I)
    for sep in ("|", "–", "—", "»"):
        if sep in t:
            partes = [p.strip() for p in t.split(sep) if p.strip()]
            if partes:
                t = max(partes[:2], key=len)
            break
    return t[:60]


CATEGORIAS_OK = {"administradoras de condominios", "sindicos profissionais"}


def _chave_cat(t):
    t = unicodedata.normalize("NFKD", str(t or ""))
    return "".join(x for x in t if not unicodedata.combining(x)).lower().strip()


def chave_nome(nome):
    """Nome reduzido a letras/numeros, sem acento, para comparar empresas.
    Nome curto demais nao serve de chave (ex.: 'Home', 'Contato')."""
    t = unicodedata.normalize("NFKD", str(nome or ""))
    t = "".join(c for c in t if not unicodedata.combining(c)).lower()
    t = re.sub(r"\b(ltda|me|epp|eireli|sa|s/a|adm|administradora|administracao|"
               r"condominios?|condominial|gestao|de|do|da|e)\b", " ", t)
    t = re.sub(r"[^a-z0-9]", "", t)
    return t if len(t) >= 4 else ""


def dominio(url):
    m = re.match(r"https?://(?:www\.)?([^/]+)", url or "")
    return m.group(1) if m else ""


def candidatos():
    """Junta as duas origens numa lista de dicts."""
    saida = []

    # --- 1) telefone declarado na ficha do diretorio ---
    if os.path.exists(FICHAS):
        for f in json.load(open(FICHAS, encoding="utf-8")).values():
            tel = normalizar(f.get("telefone"))
            if not tel:
                continue
            # categoria tem que bater EXATO: "Curso para sindicos profissionais"
            # e escola, nao e sindico
            alvo = f.get("categoria_alvo")
            if alvo is not None and not any(_chave_cat(c) in CATEGORIAS_OK for c in alvo):
                continue
            cidade = f.get("cidade") or ""
            uf = f.get("uf") or UF_POR_DDD.get(tel[2:4], "")
            nome = limpar_nome(f.get("nome")) or "(sem nome)"
            if fora_do_ramo(nome):          # gerador, elevador, advocacia...
                continue
            if GENERICO.match(nome):        # "Administrador", "Condominios"...
                nome = f"{nome} ({cidade})" if cidade else \
                       f"{nome} ({f.get('slug','')})"
            saida.append({
                "nome": nome,
                "tel": tel,
                "cidade": f"{cidade} - {uf}".strip(" -"),
                "evidencia": "ficha",
                "fonte": f.get("fonte", ""),
            })

    # --- 2) coleta nos sites das proprias empresas ---
    if os.path.exists(CSV_SITE):
        # dominios que vieram da ficha ja passaram pelo filtro de categoria;
        # os que vieram das listas de associadas (AABIC, Secovi) trazem junto
        # advogado, elevador, terceirizada... e precisam do titulo confirmando.
        # dominio -> nome da empresa na ficha. O nome da ficha e melhor que o
        # titulo da pagina, que as vezes e uma frase de propaganda.
        de_ficha = {}
        if os.path.exists(FICHAS):
            for f in json.load(open(FICHAS, encoding="utf-8")).values():
                d = dominio("https://" + re.sub(r"^https?://", "", f.get("site") or ""))
                if d:
                    de_ficha[d.lower()] = limpar_nome(f.get("nome"))

        por_pagina = {}
        for r in csv.DictReader(open(CSV_SITE, encoding="utf-8")):
            tel = normalizar(r.get("telefone"))
            if not tel or LIXO.search(r.get("nome") or ""):
                continue
            p = por_pagina.setdefault(r["fonte"], {"melhor": None, "tels": set()})
            p["tels"].add(tel)
            if p["melhor"] is None or FORCA[r["evidencia"]] > FORCA[p["melhor"]["evidencia"]]:
                p["melhor"] = {**r, "telefone": tel}

        for fonte, p in por_pagina.items():
            r = p["melhor"]
            # so numero solto no texto e mais de um numero na pagina: nao da
            # para saber qual e o da empresa (deu numero de outro estado)
            if r["evidencia"] == "texto" and len(p["tels"]) > 1:
                continue
            dom = dominio(fonte).lower()
            nome = de_ficha.get(dom) or limpar_nome(r.get("nome")) or dom
            if dom not in de_ficha and not RE_RAMO.search(nome):
                continue
            if fora_do_ramo(nome) or fora_do_ramo(r.get("nome") or ""):
                continue
            saida.append({
                "nome": nome,
                "tel": r["telefone"],
                "cidade": UF_POR_DDD.get(r["telefone"][2:4], ""),
                "evidencia": r["evidencia"],
                "fonte": fonte,
            })
    return saida


def main():
    gravar = "--gravar" in sys.argv
    todos = candidatos()

    # telefone unico, ficando com a evidencia mais forte
    vistos, finais = set(), []
    for c in sorted(todos, key=lambda x: -FORCA[x["evidencia"]]):
        if c["tel"] in vistos:
            continue
        vistos.add(c["tel"])
        finais.append(c)

    wb = load_workbook(PLANILHA)
    ws = wb.worksheets[0]
    ja, ja_nome = set(), set()
    for r in range(PRIMEIRA_LINHA, ws.max_row + 1):
        t = re.sub(r"\D", "", str(ws.cell(row=r, column=3).value or ""))
        if len(t) == 11:
            ja.add("55" + t)
        n = chave_nome(ws.cell(row=r, column=2).value)
        if n:
            ja_nome.add(n)

    # telefone repetido nunca entra; nome repetido tambem nao, para a mesma
    # empresa nao receber duas mensagens em numeros diferentes
    novos = []
    for c in finais:
        if c["tel"] in ja:
            continue
        n = chave_nome(c["nome"])
        if n and n in ja_nome:
            continue
        if n:
            ja_nome.add(n)
        novos.append(c)

    print(f"candidatos brutos : {len(todos)}")
    print(f"telefones unicos  : {len(finais)}")
    print(f"ja na planilha    : {len(finais) - len(novos)}")
    print(f"NOVOS a incluir   : {len(novos)}")
    porev = {}
    for n in novos:
        porev[n["evidencia"]] = porev.get(n["evidencia"], 0) + 1
    print("por evidencia     :", porev)
    poruf = {}
    for n in novos:
        uf = UF_POR_DDD.get(n["tel"][2:4], "?")
        poruf[uf] = poruf.get(uf, 0) + 1
    print("por UF            :", dict(sorted(poruf.items(), key=lambda x: -x[1])))

    for i, c in enumerate(novos[:40], 1):
        print(f"{i:4d}. {formatar(c['tel'])}  {c['evidencia']:6s}  {c['cidade'][:28]:28s} {c['nome'][:42]}")
    if len(novos) > 40:
        print(f"      ... e mais {len(novos)-40}")

    if not gravar:
        print("\n(nada gravado - rode com --gravar)")
        return

    bkp = PLANILHA.replace(".xlsx", f"_bkp_{datetime.now():%Y%m%d-%H%M%S}.xlsx")
    shutil.copy2(PLANILHA, bkp)
    print("\nCopia de seguranca:", bkp)

    linha = ws.max_row + 1
    seq = max((ws.cell(row=r, column=1).value or 0)
              for r in range(PRIMEIRA_LINHA, ws.max_row + 1)
              if isinstance(ws.cell(row=r, column=1).value, int))
    for c in novos:
        seq += 1
        ws.cell(row=linha, column=1, value=seq)
        ws.cell(row=linha, column=2, value=c["nome"])
        ws.cell(row=linha, column=3, value=formatar(c["tel"]))
        ws.cell(row=linha, column=4, value=c["cidade"])
        ws.cell(row=linha, column=5, value=f"Celular/WhatsApp ({c['evidencia']})")
        ws.cell(row=linha, column=6, value=c["fonte"])
        ws.cell(row=linha, column=9, value="Sindico/Administradora")
        linha += 1

    # nota do cabecalho: registra de onde veio esta leva
    nota = ws.cell(row=2, column=1).value or ""
    marca = f"Em {datetime.now():%d/%m/%Y} entraram {len(novos)} administradoras de condominio"
    if "administradoras de condominio" not in nota:
        ws.cell(row=2, column=1, value=(
            nota.rstrip() + f" {marca}, coletadas no diretorio publico de fornecedores do "
            "SindicoNet/CoteiBem e nos sites das proprias empresas; a coluna Tipo de numero "
            "traz 'ficha' quando o telefone veio do cadastro da empresa no diretorio."))

    wb.save(PLANILHA)
    print(f"Incluidos {len(novos)} contatos. Planilha: {PLANILHA}")


if __name__ == "__main__":
    main()
