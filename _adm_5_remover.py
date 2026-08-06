# -*- coding: utf-8 -*-
"""Remove da planilha os contatos passados na linha de comando e renumera.

Serve para tirar numero que nao se confirma na fonte na hora da conferencia.

Uso: python _adm_5_remover.py "(65) 99734-5198" "(21) 97949-0441"
"""
import os, re, sys, shutil
from datetime import datetime
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

BASE     = os.path.dirname(os.path.abspath(__file__))
PLANILHA = os.path.join(BASE, "Comercial", "Sindicos_Profissionais_Brasil.xlsx")
PRIMEIRA_LINHA = 4

alvos = {re.sub(r"\D", "", a) for a in sys.argv[1:] if re.sub(r"\D", "", a)}
if not alvos:
    print("informe os telefones a remover"); sys.exit(1)

wb = load_workbook(PLANILHA)
ws = wb.worksheets[0]

remover = []
for r in range(PRIMEIRA_LINHA, ws.max_row + 1):
    t = re.sub(r"\D", "", str(ws.cell(row=r, column=3).value or ""))
    if t in alvos:
        remover.append((r, ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value))

if not remover:
    print("nenhuma linha encontrada com esses telefones"); sys.exit(1)

for r, nome, tel in remover:
    print(f"  linha {r}: {tel}  {nome}")

bkp = PLANILHA.replace(".xlsx", f"_bkp_{datetime.now():%Y%m%d-%H%M%S}.xlsx")
shutil.copy2(PLANILHA, bkp)
print("Copia de seguranca:", bkp)

for r, _, _ in sorted(remover, reverse=True):
    ws.delete_rows(r)

seq = 0
for r in range(PRIMEIRA_LINHA, ws.max_row + 1):
    if ws.cell(row=r, column=3).value:
        seq += 1
        ws.cell(row=r, column=1, value=seq)

wb.save(PLANILHA)
print(f"Removidos {len(remover)} contatos. Planilha renumerada ate {seq}.")
